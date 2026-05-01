#!/usr/bin/env python3
"""Reverse-engineer a SOUP register from dependency manifests in an existing codebase.

Parses dependency manifests across Python, JS/TS, Go, and Java/Kotlin.
Produces an IEC 62304 §5.3.3 compliant XLSX with companion gap report.

Only uses openpyxl + stdlib (Python 3.11+ for tomllib).

Usage:
    python scripts/generate_soup_register.py --source /path/to/repo
    python scripts/generate_soup_register.py --source /path/to/repo --scope scope.md
    python scripts/generate_soup_register.py --example
"""

import argparse
import json
import os
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

try:
    import tomllib  # Python 3.11+
except ImportError:
    tomllib = None  # type: ignore[assignment]

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Style constants (matches design-controls pattern)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
YELLOW_FONT = Font(color="9C6500")
ORANGE_FONT = Font(color="C65911")

# License tiers
WARNING_LICENSES = {"GPL", "GPL-2.0", "GPL-3.0", "AGPL", "AGPL-3.0", "SSPL", "UNKNOWN"}
GPL_FAMILY = {"GPL", "GPL-2.0", "GPL-2.0-only", "GPL-2.0-or-later",
              "GPL-3.0", "GPL-3.0-only", "GPL-3.0-or-later",
              "LGPL-2.0", "LGPL-2.1", "LGPL-3.0",
              "AGPL-3.0", "AGPL-3.0-only", "AGPL-3.0-or-later",
              "SSPL-1.0", "SSPL"}


# ---------------------------------------------------------------------------
# Dependency data model
# ---------------------------------------------------------------------------
class Dependency:
    """A single SOUP dependency."""

    def __init__(
        self,
        name: str,
        version: str = "UNKNOWN",
        license_id: str = "UNKNOWN",
        relationship: str = "declared",
        repo_url: str = "GAP",
        last_release: str = "GAP",
        version_pinned: bool = False,
        language: str = "",
        manifest_file: str = "",
    ):
        self.name = name
        self.version = version
        self.license_id = license_id
        self.relationship = relationship
        self.repo_url = repo_url
        self.last_release = last_release
        self.version_pinned = version_pinned
        self.language = language
        self.manifest_file = manifest_file

    @property
    def license_flag(self) -> str:
        """Classify license as OK or WARNING."""
        if not self.license_id or self.license_id == "UNKNOWN":
            return "WARNING"
        normalized = self.license_id.upper().replace(" ", "-")
        for gpl in GPL_FAMILY:
            if gpl.upper() in normalized:
                return "WARNING"
        return "OK"


# ---------------------------------------------------------------------------
# Manifest parsers — stdlib only, no network calls
# ---------------------------------------------------------------------------
def parse_requirements_txt(path: Path) -> list[Dependency]:
    """Parse requirements.txt — declared deps only."""
    deps: list[Dependency] = []
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or line.startswith("-"):
            continue
        # Handle ==, >=, ~=, etc.
        match = re.match(r"^([A-Za-z0-9_.-]+)\s*(?:[><=!~]+\s*(.+?))?(?:\s*;.*)?$", line)
        if match:
            name = match.group(1)
            version = match.group(2) or "UNKNOWN"
            deps.append(Dependency(
                name=name, version=version.strip(),
                relationship="declared", manifest_file=str(path),
                language="Python",
            ))
    return deps


def parse_pipfile_lock(path: Path) -> list[Dependency]:
    """Parse Pipfile.lock — includes transitive deps with versions."""
    deps: list[Dependency] = []
    data = json.loads(path.read_text())
    for section in ("default", "develop"):
        packages = data.get(section, {})
        for name, info in packages.items():
            version = info.get("version", "UNKNOWN").lstrip("=")
            deps.append(Dependency(
                name=name, version=version,
                relationship="declared",  # Pipfile.lock flattens all deps
                version_pinned=True, manifest_file=str(path),
                language="Python",
            ))
    return deps


def parse_pyproject_toml(path: Path) -> list[Dependency]:
    """Parse pyproject.toml [project.dependencies] using tomllib."""
    if tomllib is None:
        return []
    data = tomllib.loads(path.read_text())
    deps_list = data.get("project", {}).get("dependencies", [])
    deps: list[Dependency] = []
    for dep_str in deps_list:
        match = re.match(r"^([A-Za-z0-9_.-]+)", dep_str)
        if match:
            name = match.group(1)
            version_match = re.search(r"[><=!~]+\s*([0-9][0-9A-Za-z.*-]*)", dep_str)
            version = version_match.group(1) if version_match else "UNKNOWN"
            deps.append(Dependency(
                name=name, version=version,
                relationship="declared", manifest_file=str(path),
                language="Python",
            ))
    return deps


def parse_package_lock_json(path: Path) -> list[Dependency]:
    """Parse package-lock.json v2/v3 — includes transitive deps."""
    deps: list[Dependency] = []
    data = json.loads(path.read_text())

    # v2/v3 format: "packages" key
    packages = data.get("packages", {})
    if packages:
        for pkg_path, info in packages.items():
            if not pkg_path:  # root package
                continue
            # Extract package name from path
            name = pkg_path.split("node_modules/")[-1]
            if not name:
                continue
            version = info.get("version", "UNKNOWN")
            license_id = info.get("license", "UNKNOWN")
            if isinstance(license_id, dict):
                license_id = license_id.get("type", "UNKNOWN")
            # Determine if declared or transitive
            is_dev = info.get("dev", False)
            relationship = "declared" if not info.get("_from") else "transitive"
            # Simple heuristic: top-level node_modules/ entries without nested path
            if pkg_path.count("node_modules/") == 1:
                relationship = "declared"
            elif pkg_path.count("node_modules/") > 1:
                relationship = "transitive"

            deps.append(Dependency(
                name=name, version=version, license_id=license_id,
                relationship=relationship, version_pinned=True,
                manifest_file=str(path), language="JavaScript",
            ))
    else:
        # v1 format: "dependencies" key
        for name, info in data.get("dependencies", {}).items():
            version = info.get("version", "UNKNOWN")
            deps.append(Dependency(
                name=name, version=version,
                relationship="declared", version_pinned=True,
                manifest_file=str(path), language="JavaScript",
            ))
            # Recurse for transitive
            for sub_name, sub_info in info.get("dependencies", {}).items():
                sub_version = sub_info.get("version", "UNKNOWN")
                deps.append(Dependency(
                    name=sub_name, version=sub_version,
                    relationship="transitive", version_pinned=True,
                    manifest_file=str(path), language="JavaScript",
                ))

    return deps


def parse_package_json(path: Path) -> list[Dependency]:
    """Parse package.json — declared deps only (no lock file)."""
    deps: list[Dependency] = []
    data = json.loads(path.read_text())
    for section in ("dependencies", "devDependencies"):
        for name, version_spec in data.get(section, {}).items():
            # Strip semver ranges
            version = re.sub(r"^[^0-9]*", "", version_spec) or version_spec
            deps.append(Dependency(
                name=name, version=version,
                relationship="declared", manifest_file=str(path),
                language="JavaScript",
            ))
    return deps


def parse_go_mod(path: Path) -> list[Dependency]:
    """Parse go.mod — declared deps."""
    deps: list[Dependency] = []
    in_require = False
    for line in path.read_text().splitlines():
        line = line.strip()
        if line.startswith("require ("):
            in_require = True
            continue
        if in_require and line == ")":
            in_require = False
            continue
        if in_require or line.startswith("require "):
            parts = line.replace("require ", "").strip().split()
            if len(parts) >= 2:
                name = parts[0]
                version = parts[1]
                # Skip indirect deps marker
                relationship = "transitive" if "// indirect" in line else "declared"
                deps.append(Dependency(
                    name=name, version=version,
                    relationship=relationship, manifest_file=str(path),
                    language="Go",
                ))
    return deps


def parse_go_sum(path: Path) -> list[Dependency]:
    """Parse go.sum — all resolved deps (transitive included)."""
    seen: set[str] = set()
    deps: list[Dependency] = []
    for line in path.read_text().splitlines():
        parts = line.strip().split()
        if len(parts) >= 2:
            name = parts[0]
            version = parts[1].split("/")[0]  # Remove /go.mod suffix
            key = f"{name}@{version}"
            if key not in seen:
                seen.add(key)
                deps.append(Dependency(
                    name=name, version=version,
                    relationship="transitive", version_pinned=True,
                    manifest_file=str(path), language="Go",
                ))
    return deps


def parse_pom_xml(path: Path) -> list[Dependency]:
    """Parse pom.xml <dependencies> section."""
    deps: list[Dependency] = []
    tree = ElementTree.parse(path)
    root = tree.getroot()
    # Handle Maven namespace
    ns = ""
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"

    for dep_elem in root.iter(f"{ns}dependency"):
        group = dep_elem.findtext(f"{ns}groupId", "")
        artifact = dep_elem.findtext(f"{ns}artifactId", "")
        version = dep_elem.findtext(f"{ns}version", "UNKNOWN")
        name = f"{group}:{artifact}" if group else artifact
        deps.append(Dependency(
            name=name, version=version,
            relationship="declared", manifest_file=str(path),
            language="Java",
        ))
    return deps


# ---------------------------------------------------------------------------
# Source tree scanner
# ---------------------------------------------------------------------------
MANIFEST_PARSERS: dict[str, Any] = {
    "requirements.txt": parse_requirements_txt,
    "Pipfile.lock": parse_pipfile_lock,
    "pyproject.toml": parse_pyproject_toml,
    "package-lock.json": parse_package_lock_json,
    "package.json": parse_package_json,
    "go.mod": parse_go_mod,
    "go.sum": parse_go_sum,
    "pom.xml": parse_pom_xml,
}

# Lock files — presence means versions are pinned
LOCK_FILES = {"Pipfile.lock", "package-lock.json", "yarn.lock", "pnpm-lock.yaml", "go.sum"}

# Directories to skip
SKIP_DIRS = {"node_modules", ".venv", "venv", "vendor", "dist", "build",
             "target", ".git", "__pycache__", ".tox", ".mypy_cache"}


def scan_source_tree(source: Path) -> tuple[list[Dependency], list[str], list[str]]:
    """Scan source tree for manifest files and parse dependencies.

    Returns (deps, manifests_found, gap_notes).
    """
    all_deps: list[Dependency] = []
    manifests_found: list[str] = []
    gap_notes: list[str] = []
    languages_detected: set[str] = set()
    lock_files_found: set[str] = set()
    manifest_files_found: set[str] = set()

    for root_dir, dirs, files in os.walk(source):
        # Prune skip dirs
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]

        for fname in files:
            if fname in MANIFEST_PARSERS:
                fpath = Path(root_dir) / fname
                rel_path = str(fpath.relative_to(source))
                manifests_found.append(rel_path)
                manifest_files_found.add(fname)

                if fname in LOCK_FILES:
                    lock_files_found.add(fname)

                try:
                    deps = MANIFEST_PARSERS[fname](fpath)
                    for dep in deps:
                        dep.manifest_file = rel_path
                    all_deps.extend(deps)
                    if deps:
                        languages_detected.add(deps[0].language)
                except Exception as e:
                    gap_notes.append(f"Failed to parse {rel_path}: {e}")

    # Check for missing lock files
    if "package.json" in manifest_files_found and "package-lock.json" not in lock_files_found:
        # Check for yarn.lock or pnpm-lock.yaml
        if "yarn.lock" not in manifest_files_found and "pnpm-lock.yaml" not in manifest_files_found:
            gap_notes.append(
                "Lock file missing for JavaScript/TypeScript. "
                "Dependency versions are not pinned. Resolved tree cannot be determined."
            )

    if "requirements.txt" in manifest_files_found and "Pipfile.lock" not in lock_files_found:
        gap_notes.append(
            "Lock file missing for Python (no Pipfile.lock). "
            "Transitive dependencies cannot be resolved from requirements.txt alone."
        )

    if "go.mod" in manifest_files_found and "go.sum" not in lock_files_found:
        gap_notes.append(
            "go.sum missing. Go module checksums not available. "
            "Transitive dependency verification cannot be confirmed."
        )

    # Mark pinned status based on lock files
    for dep in all_deps:
        if dep.manifest_file and Path(dep.manifest_file).name in LOCK_FILES:
            dep.version_pinned = True

    return all_deps, manifests_found, gap_notes


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------
def deduplicate_deps(deps: list[Dependency]) -> list[Dependency]:
    """Deduplicate dependencies, preferring lock-file-sourced entries."""
    seen: dict[str, Dependency] = {}
    for dep in deps:
        key = f"{dep.name}@{dep.version}"
        if key not in seen:
            seen[key] = dep
        else:
            # Prefer the one with more info (pinned > not pinned)
            if dep.version_pinned and not seen[key].version_pinned:
                seen[key] = dep
    # Sort: declared first, then transitive, then alphabetical
    result = sorted(seen.values(), key=lambda d: (
        0 if d.relationship == "declared" else 1,
        d.name.lower(),
    ))
    return result


# ---------------------------------------------------------------------------
# Example data
# ---------------------------------------------------------------------------
def get_example_deps() -> tuple[list[Dependency], list[str], list[str]]:
    """Return example dependency data for --example mode."""
    deps = [
        Dependency("flask", "3.0.2", "BSD-3-Clause", "declared", language="Python",
                    repo_url="https://github.com/pallets/flask", version_pinned=True),
        Dependency("werkzeug", "3.0.1", "BSD-3-Clause", "transitive", language="Python",
                    repo_url="https://github.com/pallets/werkzeug", version_pinned=True),
        Dependency("jinja2", "3.1.3", "BSD-3-Clause", "transitive", language="Python",
                    repo_url="https://github.com/pallets/jinja", version_pinned=True),
        Dependency("markupsafe", "2.1.5", "BSD-3-Clause", "transitive", language="Python",
                    repo_url="https://github.com/pallets/markupsafe", version_pinned=True),
        Dependency("click", "8.1.7", "BSD-3-Clause", "transitive", language="Python",
                    repo_url="https://github.com/pallets/click", version_pinned=True),
        Dependency("itsdangerous", "2.1.2", "BSD-3-Clause", "transitive", language="Python",
                    repo_url="https://github.com/pallets/itsdangerous", version_pinned=True),
        Dependency("blinker", "1.7.0", "MIT", "transitive", language="Python",
                    repo_url="https://github.com/pallets-eco/blinker", version_pinned=True),
        Dependency("numpy", "1.26.4", "BSD-3-Clause", "declared", language="Python",
                    repo_url="https://github.com/numpy/numpy", version_pinned=True),
        Dependency("pandas", "2.2.1", "BSD-3-Clause", "declared", language="Python",
                    repo_url="https://github.com/pandas-dev/pandas", version_pinned=True),
        Dependency("python-dateutil", "2.9.0", "Apache-2.0", "transitive", language="Python",
                    repo_url="https://github.com/dateutil/dateutil", version_pinned=True),
        Dependency("pytz", "2024.1", "MIT", "transitive", language="Python",
                    repo_url="https://github.com/stub42/pytz", version_pinned=True),
        Dependency("six", "1.16.0", "MIT", "transitive", language="Python",
                    repo_url="https://github.com/benjaminp/six", version_pinned=True),
        Dependency("pylint-django", "2.5.5", "GPL-2.0", "declared", language="Python",
                    repo_url="https://github.com/PyCQA/pylint-django", version_pinned=True),
    ]
    manifests = ["requirements.txt", "Pipfile.lock"]
    gaps: list[str] = []
    return deps, manifests, gaps


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------
def style_header_row(ws: Any, headers: list[str]) -> None:
    """Write and style the header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws: Any, min_width: int = 12, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                line_max = max(len(line) for line in lines)
                max_len = max(max_len, line_max)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def style_data_rows(ws: Any, start_row: int, end_row: int, num_cols: int) -> None:
    """Apply alignment and borders to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER


def build_document_control(wb: Workbook, source_name: str) -> None:
    """Build Document_Control sheet."""
    ws = wb.active
    ws.title = "Document_Control"

    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="Document Control")
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGN
    title_cell.border = THIN_BORDER
    for col in range(2, 5):
        ws.cell(row=1, column=col).border = THIN_BORDER
        ws.cell(row=1, column=col).fill = HEADER_FILL

    metadata = [
        ("Document Title", f"SOUP Register — {source_name}"),
        ("Document Version", "1.0 (draft — generated by code-to-soup-register)"),
        ("Date Generated", date.today().strftime("%Y-%m-%d")),
        ("Source Tree", source_name),
        ("Standard Reference", "IEC 62304:2006+A1:2015 §5.3.3, §8.1.2"),
    ]
    for idx, (label, value) in enumerate(metadata):
        row = 3 + idx
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = CELL_ALIGN
        label_cell.border = THIN_BORDER
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.alignment = CELL_ALIGN
        value_cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def build_soup_register(wb: Workbook, deps: list[Dependency]) -> None:
    """Build the SOUP_Register sheet."""
    ws = wb.create_sheet("SOUP_Register")
    headers = [
        "SOUP ID", "Component Name", "Version", "Declared / Transitive",
        "License", "License Flag", "Repository URL", "Last Release Date",
        "Fitness for Intended Use", "Known Anomalies Reviewed",
        "Integration Requirements", "Version Pinned", "Gap Status",
    ]
    style_header_row(ws, headers)

    for i, dep in enumerate(deps, 2):
        soup_id = f"SOUP-{i - 1:03d}"
        ws.cell(row=i, column=1, value=soup_id)
        ws.cell(row=i, column=2, value=dep.name)
        ws.cell(row=i, column=3, value=dep.version)
        ws.cell(row=i, column=4, value=dep.relationship)
        ws.cell(row=i, column=5, value=dep.license_id)
        ws.cell(row=i, column=6, value=dep.license_flag)
        ws.cell(row=i, column=7, value=dep.repo_url)
        ws.cell(row=i, column=8, value=dep.last_release)
        ws.cell(row=i, column=9, value="GAP")
        ws.cell(row=i, column=10, value="GAP")
        ws.cell(row=i, column=11, value="GAP")
        ws.cell(row=i, column=12, value="Yes" if dep.version_pinned else "No")
        # Gap Status formula
        ws.cell(row=i, column=13,
                value=f'=IF(OR(I{i}="GAP",J{i}="GAP",K{i}="GAP"),"GAP",IF(F{i}="WARNING","WARNING","COMPLETE"))')

    end_row = len(deps) + 1
    style_data_rows(ws, 2, end_row, len(headers))

    # Conditional formatting on Gap Status (column M)
    status_range = f"M2:M{end_row}"
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"COMPLETE"'], fill=GREEN_FILL, font=GREEN_FONT),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"GAP"'], fill=RED_FILL, font=RED_FONT),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"WARNING"'], fill=ORANGE_FILL, font=ORANGE_FONT),
    )

    # Conditional formatting on License Flag (column F)
    flag_range = f"F2:F{end_row}"
    ws.conditional_formatting.add(
        flag_range,
        CellIsRule(operator="equal", formula=['"WARNING"'], fill=ORANGE_FILL, font=ORANGE_FONT),
    )
    ws.conditional_formatting.add(
        flag_range,
        CellIsRule(operator="equal", formula=['"OK"'], fill=GREEN_FILL, font=GREEN_FONT),
    )

    auto_width(ws)


# ---------------------------------------------------------------------------
# Gap report generation
# ---------------------------------------------------------------------------
def generate_gap_report(
    deps: list[Dependency],
    manifests: list[str],
    gap_notes: list[str],
    source_name: str,
    scope_path: str | None,
    xlsx_path: str,
) -> str:
    """Generate companion gap report markdown."""
    today = date.today().strftime("%Y-%m-%d")
    total = len(deps)
    warning_count = sum(1 for d in deps if d.license_flag == "WARNING")
    # All deps have GAP in fitness/anomalies/integration
    gap_count = total  # Every dep has at least one GAP field
    complete_count = 0
    declared_count = sum(1 for d in deps if d.relationship == "declared")
    transitive_count = total - declared_count

    scope_ref = scope_path or "none provided"
    auth_ref = f"See {scope_path}" if scope_path else "none provided"

    lines = [
        "---",
        "type: gap-report",
        "status: draft",
        f'owner: "@[github-handle]"',
        f"generated-on: {today}",
        "generated-by: code-to-soup-register",
        f"scope-statement: {scope_ref}",
        f"source-tree: {source_name}",
        f"companion-artifact: {xlsx_path}",
        f"authorization: {auth_ref}",
        "---",
        f"# Gap Report — SOUP Register (Retrospective)",
        "",
        "## Summary",
        f"- Items identified: {total}",
        f"- Complete: {complete_count} (0%)",
        f"- Gaps: {gap_count} ({100 if total else 0}%)",
        f"- Warnings: {warning_count} (license issues)",
        f"- Declared dependencies: {declared_count}",
        f"- Transitive dependencies: {transitive_count}",
        "",
        "## Manifests Analyzed",
        "",
    ]
    for m in manifests:
        lines.append(f"- `{m}`")
    lines.append("")

    if gap_notes:
        lines.append("## Infrastructure Gaps")
        lines.append("")
        for note in gap_notes:
            lines.append(f"- {note}")
        lines.append("")

    # SOUP-specific gap categories
    lines.extend([
        "## Gaps by Category",
        "",
        f"### Evaluation Required ({total})",
        "Fitness for intended use not assessed for any dependency.",
        "Owner: Engineering.",
        "",
        "| SOUP ID | Component | Version | Action |",
        "|---------|-----------|---------|--------|",
    ])
    for i, dep in enumerate(deps, 1):
        lines.append(f"| SOUP-{i:03d} | {dep.name} | {dep.version} | Assess fitness for intended use |")

    lines.extend([
        "",
        f"### Anomaly Review Required ({total})",
        "Known issues and CVEs not reviewed for any dependency.",
        "Owner: Engineering.",
        "",
    ])

    if warning_count > 0:
        warning_deps = [d for d in deps if d.license_flag == "WARNING"]
        lines.extend([
            f"### License Review Required ({warning_count})",
            "Problematic or unknown license detected.",
            "Owner: Legal / Regulatory Affairs.",
            "",
            "| SOUP ID | Component | License | Concern |",
            "|---------|-----------|---------|---------|",
        ])
        for i, dep in enumerate(deps, 1):
            if dep.license_flag == "WARNING":
                concern = "Unknown license" if dep.license_id == "UNKNOWN" else f"Copyleft ({dep.license_id})"
                lines.append(f"| SOUP-{i:03d} | {dep.name} | {dep.license_id} | {concern} |")
        lines.append("")

    lock_missing = [n for n in gap_notes if "Lock file missing" in n or "go.sum missing" in n]
    if lock_missing:
        lines.extend([
            f"### Lock File Missing ({len(lock_missing)})",
            "Dependency versions not pinned. Transitive tree unresolvable.",
            "Owner: Engineering.",
            "",
        ])
        for note in lock_missing:
            lines.append(f"- {note}")
        lines.append("")

    lines.extend([
        f"### Integration Documentation Required ({total})",
        "SOUP integration requirements not documented for any dependency.",
        "Owner: Engineering.",
        "",
        "## Recommended Sequence",
        "1. **Lock files first** — pin all dependency versions with lock files",
        "2. **License review** — have legal review WARNING-flagged licenses",
        "3. **Anomaly review** — check CVE databases for each component",
        "4. **Fitness assessment** — evaluate each SOUP item for intended use suitability",
        "5. **Integration docs** — document integration requirements and constraints",
        "",
        "## What This Report Is Not",
        "This report identifies gaps in SOUP documentation per IEC 62304:2006+A1:2015.",
        "It does not constitute a SOUP evaluation, a security assessment, or any other",
        "regulated activity. The SOUP register and this gap report are inputs to a",
        "human-led retrospective compliance effort, not substitutes for it.",
    ])

    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def slugify(name: str) -> str:
    """Convert name to filename-safe slug."""
    slug = name.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    return slug.strip("-")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Reverse-engineer a SOUP register from dependency manifests (XLSX + gap report)."
    )
    parser.add_argument(
        "--source", type=str,
        help="Path to source tree root.",
    )
    parser.add_argument(
        "--scope", type=str, default=None,
        help="Path to scope statement (optional, recommended).",
    )
    parser.add_argument(
        "--output-dir", type=str, default=None,
        help="Directory for output files (default: output/ in repo root).",
    )
    parser.add_argument(
        "--example", action="store_true",
        help="Generate example output (no source tree needed).",
    )
    args = parser.parse_args()

    if args.example:
        deps, manifests, gap_notes = get_example_deps()
        source_name = "example-samd-app"
    elif args.source:
        source = Path(args.source).resolve()
        if not source.is_dir():
            print(f"Error: {args.source} is not a directory.", file=sys.stderr)
            sys.exit(1)
        source_name = source.name
        deps, manifests, gap_notes = scan_source_tree(source)
        if not deps:
            print(f"No dependency manifests found in {args.source}.", file=sys.stderr)
            print("Supported: requirements.txt, package-lock.json, go.mod, pom.xml, etc.")
            sys.exit(1)
    else:
        parser.error("Provide --source /path/to/repo or --example.")

    # Deduplicate
    deps = deduplicate_deps(deps)

    if not args.scope and not args.example:
        print("WARNING: No scope statement provided. Gap report will be flagged by "
              "qa-reviewer as lacking procedural authorization.")

    # Generate XLSX
    wb = Workbook()
    build_document_control(wb, source_name)
    build_soup_register(wb, deps)
    wb.active = 0

    # TODO: Output dir resolution is duplicated in 3 scripts. Extract to a shared
    # module if/when these scripts become a package.
    # See also: code-to-design-inputs/scripts/generate_design_inputs.py
    # See also: code-to-hazard-candidates/scripts/generate_hazard_candidates.py
    if args.output_dir:
        output_dir = Path(args.output_dir).resolve()
    else:
        script_dir = Path(__file__).parent
        output_dir = script_dir.parent.parent.parent / "output"
    output_dir.mkdir(exist_ok=True)

    slug = slugify(source_name)
    xlsx_path = output_dir / f"soup-register-{slug}.xlsx"
    wb.save(str(xlsx_path))

    # Generate gap report
    gap_report_path = output_dir / f"gap-report-soup-{date.today().strftime('%Y-%m-%d')}.md"
    gap_report_content = generate_gap_report(
        deps, manifests, gap_notes, source_name, args.scope, str(xlsx_path.name),
    )
    gap_report_path.write_text(gap_report_content)

    # Summary
    declared = sum(1 for d in deps if d.relationship == "declared")
    transitive = len(deps) - declared
    warnings = sum(1 for d in deps if d.license_flag == "WARNING")

    print(f"Generated: {xlsx_path}")
    print(f"Gap report: {gap_report_path}")
    print(f"  Components: {len(deps)} ({declared} declared, {transitive} transitive)")
    print(f"  License warnings: {warnings}")
    print(f"  Manifests parsed: {len(manifests)}")
    if gap_notes:
        print(f"  Infrastructure gaps: {len(gap_notes)}")
        for note in gap_notes:
            print(f"    - {note}")


if __name__ == "__main__":
    main()
