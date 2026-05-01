#!/usr/bin/env python3
"""Reverse-engineer hazard candidates from source code in an existing codebase.

Identifies safety-critical code regions using 6 heuristic scanners and proposes
candidate hazard scenarios with failure modes. Produces an XLSX hazard candidate
list with companion gap report including coverage disclaimer.

Only uses openpyxl + stdlib (Python 3.11+).

Usage:
    python scripts/generate_hazard_candidates.py --source /path/to/repo
    python scripts/generate_hazard_candidates.py --source /path/to/repo --risk-file risk.md
    python scripts/generate_hazard_candidates.py --example
"""

import argparse
import json
import os
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants (matches design-controls / code-to-soup-register pattern)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
YELLOW_FONT = Font(color="9C6500")
ORANGE_FONT = Font(color="C65911")

# Directories to skip — broader than design-inputs per spec
HAZARD_SKIP_DIRS = {
    "node_modules", ".venv", "venv", "vendor", "dist", "build", "target",
    ".git", "__pycache__", ".tox", ".mypy_cache", ".pytest_cache",
    "test", "tests", "__tests__", ".github", "docs",
    "terraform", "k8s",
}

# File extensions to scan
SOURCE_EXTENSIONS = {
    ".py", ".js", ".ts", ".jsx", ".tsx", ".go", ".java", ".kt", ".kts",
    ".rb", ".rs", ".c", ".cpp", ".h", ".hpp", ".cs",
}


# ---------------------------------------------------------------------------
# Hazard Candidate data model
# ---------------------------------------------------------------------------
class HazardCandidate:
    """A single candidate hazard extracted from code."""

    def __init__(
        self,
        hc_id: str,
        source_file: str,
        start_line: int,
        end_line: int,
        heuristic_matched: str,
        proposed_hazard: str,
        proposed_failure_mode: str,
        proposed_harm: str,
        existing_haz_id: str = "",
        status: str = "CANDIDATE",
        dispositioning: str = "GAP",
        clinical_rationale: str = "GAP",
    ):
        self.hc_id = hc_id
        self.source_file = source_file
        self.start_line = start_line
        self.end_line = end_line
        self.heuristic_matched = heuristic_matched
        self.proposed_hazard = proposed_hazard
        self.proposed_failure_mode = proposed_failure_mode
        self.proposed_harm = proposed_harm
        self.existing_haz_id = existing_haz_id
        self.status = status
        self.dispositioning = dispositioning
        self.clinical_rationale = clinical_rationale

    @property
    def code_region(self) -> str:
        return f"{self.source_file}:{self.start_line}-{self.end_line}"


# ---------------------------------------------------------------------------
# Function boundary detection — expand trigger line to enclosing function
# ---------------------------------------------------------------------------
def _find_enclosing_function(lines: list[str], trigger_line: int) -> tuple[int, int]:
    """Find the start and end of the enclosing function for a trigger line.

    Returns (start_line, end_line) as 1-indexed line numbers.
    Falls back to trigger_line +/- 5 if no function boundary found.
    """
    # Search backward for function definition
    start = trigger_line
    for i in range(trigger_line - 1, max(0, trigger_line - 50), -1):
        line = lines[i].strip() if i < len(lines) else ""
        if re.match(r'^\s*(def |function |func |public |private |protected |static |async\s+)', line):
            start = i + 1  # 1-indexed
            break
        # Class-level method (indented def)
        if re.match(r'^\s+def\s+', line):
            start = i + 1
            break

    # Search forward for next function definition or significant outdent
    end = trigger_line
    if start < trigger_line:
        # We found a function — find its end
        start_indent = len(lines[start - 1]) - len(lines[start - 1].lstrip()) if start - 1 < len(lines) else 0
        for i in range(trigger_line, min(len(lines), trigger_line + 100)):
            line = lines[i]
            if line.strip() == "":
                continue
            current_indent = len(line) - len(line.lstrip())
            # Next function at same or less indent
            if current_indent <= start_indent and re.match(r'\s*(def |function |func |class |public |private )', line.strip()):
                end = i  # 1-indexed (exclusive, so this line is not included)
                break
            end = i + 1
    else:
        # No function found — use a window around trigger line
        start = max(1, trigger_line - 5)
        end = min(len(lines), trigger_line + 5)

    return start, end


# ---------------------------------------------------------------------------
# Heuristic scanner functions
# ---------------------------------------------------------------------------
def scan_alarm_logic(content: str, filepath: str) -> list[HazardCandidate]:
    """Detect alarm state evaluation, suppression, escalation logic."""
    results: list[HazardCandidate] = []
    lines = content.splitlines()
    patterns = [
        (r'(?:alarm|alert)\s*[=<>!]+|trigger_alarm|fire_alarm|raise_alarm',
         "Alarm may fail to trigger",
         "Missed alarm — condition met but alarm not fired",
         "Delayed or absent clinical intervention due to missed alarm"),
        (r'(?:suppress|silence|mute|inhibit|snooze|acknowledge).*(?:alarm|alert)',
         "Alarm suppression may prevent valid alarm",
         "Alarm suppression failure — valid alarm silenced",
         "Clinical condition undetected during suppression window"),
        (r'(?:alarm_state|alert_state|alarm_level|severity)\s*[=:]',
         "Alarm state transition may be incorrect",
         "Incorrect alarm escalation/de-escalation",
         "Patient receives wrong urgency level of clinical response"),
        (r'(?:escalat|de-escalat|upgrade|downgrade).*(?:alarm|alert|severity)',
         "Alarm escalation logic may malfunction",
         "Delayed alarm escalation",
         "Critical condition treated as lower priority"),
    ]

    for line_num, line in enumerate(lines, 1):
        for pattern, hazard, failure_mode, harm in patterns:
            if re.search(pattern, line, re.IGNORECASE):
                start, end = _find_enclosing_function(lines, line_num)
                results.append(HazardCandidate(
                    hc_id="",
                    source_file=filepath,
                    start_line=start,
                    end_line=end,
                    heuristic_matched="alarm_logic",
                    proposed_hazard=hazard,
                    proposed_failure_mode=failure_mode,
                    proposed_harm=harm,
                ))
    return results


def scan_threshold_calculations(content: str, filepath: str) -> list[HazardCandidate]:
    """Detect clinical threshold comparisons, dosing calculations, boundary values."""
    results: list[HazardCandidate] = []
    lines = content.splitlines()
    patterns = [
        (r'(?:spo2|heart_rate|hr|bp|temperature|dose|weight)\s*[<>=!]+\s*\d+',
         "Clinical threshold may be incorrect",
         "Incorrect threshold — value too high or too low",
         "Patient harm due to delayed or premature clinical action"),
        (r'(?:dose|dosage|mg_per_kg|ml_per_kg|units_per_kg|infusion_rate)\s*[=*]',
         "Dosing calculation may produce incorrect result",
         "Dosing miscalculation — unit conversion or weight-based error",
         "Patient receives incorrect medication dose"),
        (r'(?:convert|to_mg|to_ml|kg_to_lb|celsius_to|fahrenheit_to)',
         "Unit conversion may introduce error",
         "Calculation overflow/underflow in unit conversion",
         "Incorrect clinical value due to conversion error"),
        (r'(?:min_dose|max_dose|min_rate|max_rate|lower_limit|upper_limit)\s*[=<>]',
         "Boundary limit may not be enforced correctly",
         "Boundary violation — value exceeds safe limits",
         "Patient exposed to unsafe parameter values"),
    ]

    for line_num, line in enumerate(lines, 1):
        for pattern, hazard, failure_mode, harm in patterns:
            if re.search(pattern, line, re.IGNORECASE):
                start, end = _find_enclosing_function(lines, line_num)
                results.append(HazardCandidate(
                    hc_id="",
                    source_file=filepath,
                    start_line=start,
                    end_line=end,
                    heuristic_matched="threshold_calculations",
                    proposed_hazard=hazard,
                    proposed_failure_mode=failure_mode,
                    proposed_harm=harm,
                ))
    return results


def scan_decision_support(content: str, filepath: str) -> list[HazardCandidate]:
    """Detect classification, triage, scoring, prediction logic."""
    results: list[HazardCandidate] = []
    lines = content.splitlines()
    patterns = [
        (r'def\s+(?:classify|triage|score|predict|recommend|assess|grade|stratify|diagnose)',
         "Decision support output may be incorrect",
         "Incorrect classification — patient misclassified",
         "Patient receives inappropriate level of care based on wrong classification"),
        (r'(?:risk_score|severity_score|acuity|priority_level|triage_level)\s*[=:]',
         "Risk score computation may produce wrong result",
         "Model inconsistency — score contradicts clinical context",
         "Clinical decision based on unreliable risk assessment"),
        (r'(?:model\.predict|classifier\.predict|inference|forward_pass)',
         "ML model inference may produce incorrect output",
         "Stale or biased model output",
         "Clinical action taken based on unreliable prediction"),
    ]

    for line_num, line in enumerate(lines, 1):
        for pattern, hazard, failure_mode, harm in patterns:
            if re.search(pattern, line, re.IGNORECASE):
                start, end = _find_enclosing_function(lines, line_num)
                results.append(HazardCandidate(
                    hc_id="",
                    source_file=filepath,
                    start_line=start,
                    end_line=end,
                    heuristic_matched="decision_support",
                    proposed_hazard=hazard,
                    proposed_failure_mode=failure_mode,
                    proposed_harm=harm,
                ))
    return results


def scan_ehr_writes(content: str, filepath: str) -> list[HazardCandidate]:
    """Detect data written to EHR/clinical record."""
    results: list[HazardCandidate] = []
    lines = content.splitlines()
    patterns = [
        (r'(?:from|import)\s+(?:fhir|hl7)|(?:fhir|hl7)[\._]\w+',
         "Data written to clinical record may be incorrect",
         "Incorrect data written — wrong value or wrong patient",
         "Clinical decisions based on erroneous record data"),
        # Patient included deliberately — constructing patient resources is a hazard surface
        # Order: Bundle|Patient|Observation|DiagnosticReport|MedicationRequest (canonical)
        (r'(?:Bundle|Patient|Observation|DiagnosticReport|MedicationRequest)\s*\(',
         "Data written to clinical record may be incorrect",
         "Incorrect data written — FHIR resource constructed with wrong data",
         "Clinical decisions based on erroneous record data"),
        (r'(?:ehr|emr|chart|record)\.(?:write|create|update|post|submit|save)',
         "EHR write may fail silently",
         "Data loss — observation not persisted to clinical record",
         "Missing data leads to incomplete clinical picture"),
        (r'(?:place_order|submit_order|create_order|order_entry)',
         "Clinical order may be placed incorrectly",
         "Incorrect order — wrong medication, dose, or patient",
         "Patient receives wrong treatment based on erroneous order"),
    ]

    for line_num, line in enumerate(lines, 1):
        for pattern, hazard, failure_mode, harm in patterns:
            if re.search(pattern, line, re.IGNORECASE):
                start, end = _find_enclosing_function(lines, line_num)
                results.append(HazardCandidate(
                    hc_id="",
                    source_file=filepath,
                    start_line=start,
                    end_line=end,
                    heuristic_matched="ehr_writes",
                    proposed_hazard=hazard,
                    proposed_failure_mode=failure_mode,
                    proposed_harm=harm,
                ))
    return results


def scan_device_control(content: str, filepath: str) -> list[HazardCandidate]:
    """Detect actuator commands, pump rates, ventilator settings."""
    results: list[HazardCandidate] = []
    lines = content.splitlines()
    patterns = [
        (r'(?:set_rate|set_flow|set_pressure|set_speed|set_power|actuate|activate_pump)',
         "Device control command may set incorrect parameter",
         "Incorrect setting — wrong rate, volume, or pressure",
         "Patient exposed to unsafe device output"),
        (r'(?:tidal_volume|peep|fio2|respiratory_rate|inspiratory_pressure|pip|map_pressure)',
         "Ventilator parameter may be incorrect",
         "Runaway output — control loop failure",
         "Patient receives unsafe ventilation parameters"),
        (r'(?:infusion_rate|bolus|prime_line|pump_speed|flow_rate)',
         "Infusion control may malfunction",
         "Pump delivers incorrect flow rate",
         "Patient receives wrong infusion volume"),
        (r'(?:start_therapy|stop_therapy|pause_therapy|resume_therapy|emergency_stop)',
         "Therapy state transition may fail",
         "Command not executed — device fails to respond",
         "Therapy continues or stops inappropriately"),
    ]

    for line_num, line in enumerate(lines, 1):
        for pattern, hazard, failure_mode, harm in patterns:
            if re.search(pattern, line, re.IGNORECASE):
                start, end = _find_enclosing_function(lines, line_num)
                results.append(HazardCandidate(
                    hc_id="",
                    source_file=filepath,
                    start_line=start,
                    end_line=end,
                    heuristic_matched="device_control",
                    proposed_hazard=hazard,
                    proposed_failure_mode=failure_mode,
                    proposed_harm=harm,
                ))
    return results


def scan_failsafe_paths(content: str, filepath: str) -> list[HazardCandidate]:
    """Detect fallback behavior, degraded mode, safe state transitions."""
    results: list[HazardCandidate] = []
    lines = content.splitlines()
    patterns = [
        (r'(?:safe_state|safe_mode|degraded_mode|fallback|fail_safe|emergency_mode)',
         "Safe state transition may fail",
         "Failed transition to safe state on error",
         "System remains in unsafe state during fault condition"),
        (r'(?:watchdog|heartbeat|keepalive|timeout_handler)\s*[=:(]',
         "Watchdog/heartbeat mechanism may fail",
         "Silent failure — error without notification or safe state",
         "System failure undetected, no clinical response triggered"),
        (r'(?:degrade|fallback_to|switch_to_manual|disable_auto|revert_to_default)',
         "Graceful degradation may be incomplete",
         "Incomplete fallback — partial state transition",
         "System in undefined state with unpredictable behavior"),
    ]

    for line_num, line in enumerate(lines, 1):
        for pattern, hazard, failure_mode, harm in patterns:
            if re.search(pattern, line, re.IGNORECASE):
                start, end = _find_enclosing_function(lines, line_num)
                results.append(HazardCandidate(
                    hc_id="",
                    source_file=filepath,
                    start_line=start,
                    end_line=end,
                    heuristic_matched="failsafe_paths",
                    proposed_hazard=hazard,
                    proposed_failure_mode=failure_mode,
                    proposed_harm=harm,
                ))
    return results


# ---------------------------------------------------------------------------
# Heuristic config loader — domain-specific pattern overrides
# ---------------------------------------------------------------------------
# Maps scanner function names to their heuristic_matched label
_SCANNER_NAME_MAP = {
    "alarm_logic": scan_alarm_logic,
    "threshold_calculations": scan_threshold_calculations,
    "decision_support": scan_decision_support,
    "ehr_writes": scan_ehr_writes,
    "device_control": scan_device_control,
    "failsafe_paths": scan_failsafe_paths,
}


def _make_extra_scanner(
    heuristic_name: str,
    patterns: list[dict[str, str]],
) -> Any:
    """Create a scanner function from JSON-defined patterns."""
    compiled: list[tuple[str, str, str, str]] = []
    for p in patterns:
        compiled.append((
            p["regex"],
            p["proposed_hazard"],
            p["proposed_failure_mode"],
            p["proposed_harm"],
        ))

    def scanner(content: str, filepath: str) -> list[HazardCandidate]:
        results: list[HazardCandidate] = []
        lines = content.splitlines()
        for line_num, line in enumerate(lines, 1):
            for pattern, hazard, failure_mode, harm in compiled:
                if re.search(pattern, line, re.IGNORECASE):
                    start, end = _find_enclosing_function(lines, line_num)
                    results.append(HazardCandidate(
                        hc_id="",
                        source_file=filepath,
                        start_line=start,
                        end_line=end,
                        heuristic_matched=heuristic_name,
                        proposed_hazard=hazard,
                        proposed_failure_mode=failure_mode,
                        proposed_harm=harm,
                    ))
        return results

    return scanner


def load_heuristic_config(
    config_path: Path,
) -> tuple[list[Any], str]:
    """Load heuristic config JSON and return (scanners_list, domain_name).

    Modes per scanner:
    - extend: default scanner + extra patterns from config
    - replace: only config patterns (default scanner skipped)
    - disable: scanner skipped entirely
    Scanners not mentioned in config use defaults unchanged.
    """
    data = json.loads(config_path.read_text())
    domain = data.get("domain", "custom")
    scanner_configs = data.get("scanners", {})

    # Validate scanner config structure
    valid_modes = {"extend", "replace", "disable"}
    required_pattern_keys = {"regex", "proposed_hazard", "proposed_failure_mode", "proposed_harm"}
    for scanner_name, cfg in scanner_configs.items():
        mode = cfg.get("mode", "extend")
        if mode not in valid_modes:
            raise ValueError(
                f"Scanner '{scanner_name}': invalid mode '{mode}'. "
                f"Must be one of: {', '.join(sorted(valid_modes))}"
            )
        for i, pattern in enumerate(cfg.get("patterns", [])):
            missing = required_pattern_keys - set(pattern.keys())
            if missing:
                raise ValueError(
                    f"Scanner '{scanner_name}', pattern {i}: missing required "
                    f"keys: {', '.join(sorted(missing))}"
                )

    scanners: list[Any] = []
    for name, default_fn in _SCANNER_NAME_MAP.items():
        if name not in scanner_configs:
            scanners.append(default_fn)
            continue

        cfg = scanner_configs[name]
        mode = cfg.get("mode", "extend")

        if mode == "disable":
            continue
        elif mode == "replace":
            patterns = cfg.get("patterns", [])
            if patterns:
                scanners.append(_make_extra_scanner(name, patterns))
        else:  # extend
            scanners.append(default_fn)
            patterns = cfg.get("patterns", [])
            if patterns:
                scanners.append(_make_extra_scanner(name, patterns))

    return scanners, domain


# ---------------------------------------------------------------------------
# Source tree walker
# ---------------------------------------------------------------------------
ALL_SCANNERS = [
    scan_alarm_logic,
    scan_threshold_calculations,
    scan_decision_support,
    scan_ehr_writes,
    scan_device_control,
    scan_failsafe_paths,
]


def walk_source_tree(
    source: Path,
    scanners: list[Any] | None = None,
) -> tuple[list[HazardCandidate], list[str], list[str]]:
    """Walk source tree and run all scanners on each file.

    Returns (hazard_candidates, files_analyzed, files_excluded).
    """
    active_scanners = ALL_SCANNERS if scanners is None else scanners
    all_hcs: list[HazardCandidate] = []
    files_analyzed: list[str] = []
    files_excluded: list[str] = []

    for root_dir, dirs, files in os.walk(source):
        # Track excluded dirs
        excluded = [d for d in dirs if d in HAZARD_SKIP_DIRS]
        for d in excluded:
            files_excluded.append(str(Path(root_dir, d).relative_to(source)) + "/")

        # Prune skip dirs
        dirs[:] = [d for d in dirs if d not in HAZARD_SKIP_DIRS]

        for fname in files:
            fpath = Path(root_dir) / fname
            ext = fpath.suffix.lower()
            if ext not in SOURCE_EXTENSIONS:
                continue

            rel_path = str(fpath.relative_to(source))
            files_analyzed.append(rel_path)

            try:
                content = fpath.read_text(errors="replace")
            except (OSError, UnicodeDecodeError):
                continue

            for scanner in active_scanners:
                results = scanner(content, rel_path)
                all_hcs.extend(results)

    return all_hcs, files_analyzed, files_excluded


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------
def deduplicate_hcs(hcs: list[HazardCandidate]) -> list[HazardCandidate]:
    """Remove duplicate HCs based on overlapping code regions + same heuristic.

    After function boundary expansion, multiple trigger lines in the same
    function produce near-duplicate HCs. Collapse by (file, start, end, heuristic)
    and keep the first match.
    """
    seen: set[str] = set()
    unique: list[HazardCandidate] = []
    for hc in hcs:
        key = f"{hc.source_file}:{hc.start_line}:{hc.end_line}:{hc.heuristic_matched}"
        if key not in seen:
            seen.add(key)
            unique.append(hc)
    # Sort by file then start line
    unique.sort(key=lambda h: (h.source_file, h.start_line))
    return unique


# ---------------------------------------------------------------------------
# Risk file cross-reference
# ---------------------------------------------------------------------------
def _parse_risk_file(risk_path: Path) -> list[dict[str, str]]:
    """Parse existing risk file for HAZ-xxx IDs and associated code references.

    Returns list of {haz_id, description, code_ref}.
    """
    hazards: list[dict[str, str]] = []
    content = risk_path.read_text(errors="replace")

    for line_num, line in enumerate(content.splitlines(), 1):
        # Look for HAZ-xxx pattern
        match = re.search(r'(HAZ-\d+)', line)
        if match:
            haz_id = match.group(1)
            # Try to find file:line reference on same line
            code_ref_match = re.search(r'([a-zA-Z0-9_/.-]+\.(py|js|ts|go|java|kt)):(\d+)', line)
            code_ref = code_ref_match.group(0) if code_ref_match else ""
            hazards.append({
                "haz_id": haz_id,
                "description": line.strip()[:120],
                "code_ref": code_ref,
            })

    return hazards


def cross_reference_risk_file(
    hcs: list[HazardCandidate], risk_path: Path
) -> list[dict[str, str]]:
    """Cross-reference HCs against existing risk file.

    Updates hc.existing_haz_id for matches. Returns unmapped hazards
    (in risk file but not matched to any HC code region).
    """
    existing_hazards = _parse_risk_file(risk_path)
    if not existing_hazards:
        return []

    matched_haz_ids: set[str] = set()

    for hc in hcs:
        for hazard in existing_hazards:
            if hazard["code_ref"]:
                # Parse file:line from code_ref
                ref_match = re.match(r'(.+):(\d+)', hazard["code_ref"])
                if ref_match:
                    ref_file = ref_match.group(1)
                    ref_line = int(ref_match.group(2))
                    if hc.source_file == ref_file and hc.start_line <= ref_line <= hc.end_line:
                        hc.existing_haz_id = hazard["haz_id"]
                        matched_haz_ids.add(hazard["haz_id"])

    # Find unmapped hazards
    unmapped = [
        h for h in existing_hazards
        if h["haz_id"] not in matched_haz_ids
    ]
    return unmapped


# ---------------------------------------------------------------------------
# Example data
# ---------------------------------------------------------------------------
def get_example_data() -> tuple[list[HazardCandidate], list[str], list[str]]:
    """Return example hazard candidate data for --example mode."""
    hcs = [
        HazardCandidate("", "monitor/alarms.py", 15, 35,
                         "alarm_logic",
                         "Alarm may fail to trigger",
                         "Missed alarm — condition met but alarm not fired",
                         "Delayed clinical intervention due to missed alarm"),
        HazardCandidate("", "monitor/alarms.py", 40, 55,
                         "alarm_logic",
                         "Alarm suppression may prevent valid alarm",
                         "Alarm suppression failure — valid alarm silenced",
                         "Clinical condition undetected during suppression window"),
        HazardCandidate("", "clinical/dosing.py", 8, 25,
                         "threshold_calculations",
                         "Dosing calculation may produce incorrect result",
                         "Dosing miscalculation — weight-based error",
                         "Patient receives incorrect medication dose"),
        HazardCandidate("", "clinical/dosing.py", 30, 45,
                         "threshold_calculations",
                         "Clinical threshold may be incorrect",
                         "Incorrect threshold — value too high or too low",
                         "Patient harm due to delayed clinical action"),
        HazardCandidate("", "integrations/ehr.py", 20, 40,
                         "ehr_writes",
                         "Data written to clinical record may be incorrect",
                         "Incorrect data written — wrong value or patient",
                         "Clinical decisions based on erroneous record data"),
        HazardCandidate("", "core/safe_mode.py", 10, 30,
                         "failsafe_paths",
                         "Safe state transition may fail",
                         "Failed transition to safe state on error",
                         "System remains in unsafe state during fault"),
    ]
    files = [
        "monitor/alarms.py", "clinical/dosing.py",
        "integrations/ehr.py", "core/safe_mode.py",
    ]
    excluded = ["test/", "docs/", "node_modules/"]
    return hcs, files, excluded


# ---------------------------------------------------------------------------
# XLSX generation helpers
# ---------------------------------------------------------------------------
def style_header_row(ws: Any, headers: list[str]) -> None:
    """Write and style the header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws: Any, min_width: int = 12, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                cell_lines = str(cell.value).split("\n")
                line_max = max(len(line) for line in cell_lines)
                max_len = max(max_len, line_max)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def style_data_rows(ws: Any, start_row: int, end_row: int, num_cols: int) -> None:
    """Apply alignment and borders to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# XLSX builder
# ---------------------------------------------------------------------------
def build_document_control(wb: Workbook, source_name: str) -> None:
    """Build Document_Control sheet."""
    ws = wb.active
    ws.title = "Document_Control"

    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="Document Control")
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGN
    title_cell.border = THIN_BORDER
    for col in range(2, 5):
        ws.cell(row=1, column=col).border = THIN_BORDER
        ws.cell(row=1, column=col).fill = HEADER_FILL

    metadata = [
        ("Document Title", f"Hazard Candidate List — {source_name}"),
        ("Document Version", "1.0 (draft — generated by code-to-hazard-candidates)"),
        ("Date Generated", date.today().strftime("%Y-%m-%d")),
        ("Source Tree", source_name),
        ("Standard Reference", "ISO 14971:2019, IEC 62304:2006+A1:2015"),
    ]
    for idx, (label, value) in enumerate(metadata):
        row = 3 + idx
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = CELL_ALIGN
        label_cell.border = THIN_BORDER
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.alignment = CELL_ALIGN
        value_cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def build_hazard_candidates_sheet(wb: Workbook, hcs: list[HazardCandidate]) -> None:
    """Build the Hazard_Candidates sheet."""
    ws = wb.create_sheet("Hazard_Candidates")
    headers = [
        "Candidate ID", "Code Region", "Heuristic Matched",
        "Proposed Hazard", "Proposed Failure Mode", "Proposed Harm",
        "Existing HAZ ID", "Status", "Dispositioning Decision",
        "Clinical Rationale",
    ]
    style_header_row(ws, headers)

    for i, hc in enumerate(hcs, 2):
        ws.cell(row=i, column=1, value=hc.hc_id)
        ws.cell(row=i, column=2, value=hc.code_region)
        ws.cell(row=i, column=3, value=hc.heuristic_matched)
        ws.cell(row=i, column=4, value=hc.proposed_hazard)
        ws.cell(row=i, column=5, value=hc.proposed_failure_mode)
        ws.cell(row=i, column=6, value=hc.proposed_harm)
        ws.cell(row=i, column=7, value=hc.existing_haz_id)
        ws.cell(row=i, column=8, value=hc.status)
        ws.cell(row=i, column=9, value=hc.dispositioning)
        ws.cell(row=i, column=10, value=hc.clinical_rationale)

    end_row = len(hcs) + 1
    num_cols = len(headers)
    if end_row >= 2:
        style_data_rows(ws, 2, end_row, num_cols)

        # Conditional formatting on Status column (H)
        status_range = f"H2:H{end_row}"
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=['"CANDIDATE"'], fill=YELLOW_FILL, font=YELLOW_FONT),
        )

        # Conditional formatting on Dispositioning column (I)
        disp_range = f"I2:I{end_row}"
        ws.conditional_formatting.add(
            disp_range,
            CellIsRule(operator="equal", formula=['"GAP"'], fill=RED_FILL, font=RED_FONT),
        )

    auto_width(ws)


# ---------------------------------------------------------------------------
# Gap report generation
# ---------------------------------------------------------------------------
def generate_gap_report(
    hcs: list[HazardCandidate],
    files_analyzed: list[str],
    files_excluded: list[str],
    source_name: str,
    scope_path: str | None,
    xlsx_path: str,
    unmapped_hazards: list[dict[str, str]] | None = None,
    risk_file_path: str | None = None,
) -> str:
    """Generate companion gap report markdown with coverage disclaimer."""
    today = date.today().strftime("%Y-%m-%d")
    total = len(hcs)
    scope_ref = scope_path or "none provided"
    auth_ref = f"See {scope_path}" if scope_path else "none provided"

    # Count by heuristic
    heuristic_counts: dict[str, int] = {}
    for hc in hcs:
        heuristic_counts[hc.heuristic_matched] = heuristic_counts.get(hc.heuristic_matched, 0) + 1

    matched_count = sum(1 for hc in hcs if hc.existing_haz_id)
    new_count = total - matched_count

    lines = [
        "---",
        "type: gap-report",
        "status: draft",
        f'owner: "@[github-handle]"',
        f"generated-on: {today}",
        "generated-by: code-to-hazard-candidates",
        f"scope-statement: {scope_ref}",
        f"source-tree: {source_name}",
        f"companion-artifact: {xlsx_path}",
        f"authorization: {auth_ref}",
        "---",
        "",
        "# Gap Report — Hazard Candidates (Retrospective)",
        "",
        "## Summary",
        f"- Hazard candidates identified: {total}",
        f"- All candidates: Status = CANDIDATE, Dispositioning = GAP, Clinical Rationale = GAP",
        f"- Source files analyzed: {len(files_analyzed)}",
        f"- Directories excluded: {len(files_excluded)}",
        "",
    ]

    if risk_file_path:
        lines.extend([
            "### Risk File Cross-Reference",
            f"- Risk file: `{risk_file_path}`",
            f"- Matched to existing hazards: {matched_count}",
            f"- New candidates (not in risk file): {new_count}",
            "",
        ])

    lines.extend([
        "### By Heuristic",
    ])
    for heuristic, count in sorted(heuristic_counts.items()):
        lines.append(f"- {heuristic}: {count}")
    lines.append("")

    # Candidate details
    if hcs:
        lines.extend([
            "## Hazard Candidates",
            "",
            "| HC ID | Code Region | Heuristic | Proposed Hazard | Existing HAZ |",
            "|-------|------------|-----------|----------------|-------------|",
        ])
        for hc in hcs:
            existing = hc.existing_haz_id or "—"
            lines.append(f"| {hc.hc_id} | {hc.code_region} | {hc.heuristic_matched} | {hc.proposed_hazard[:50]} | {existing} |")
        lines.append("")

    # Unmapped hazards from risk file
    unmapped_hazards = unmapped_hazards or []
    if unmapped_hazards:
        lines.extend([
            f"## Unmapped Existing Hazards ({len(unmapped_hazards)})",
            "These hazards exist in the risk file but do not map to any identified code region.",
            "They may be infrastructure-level, operational, or use-environment hazards.",
            "",
            "| HAZ ID | Description |",
            "|--------|-------------|",
        ])
        for h in unmapped_hazards:
            lines.append(f"| {h['haz_id']} | {h['description'][:80]} |")
        lines.append("")

    # Coverage section (mandatory per spec)
    lines.extend([
        "## Coverage",
        "",
        "### Code paths analyzed",
    ])
    for f in sorted(files_analyzed):
        lines.append(f"- `{f}`")
    lines.extend([
        f"- Heuristics applied: {', '.join(sorted(set(hc.heuristic_matched for hc in hcs))) if hcs else 'all 6 (no matches)'}",
        "",
        "### Code paths excluded",
    ])
    if files_excluded:
        for f in sorted(set(files_excluded)):
            lines.append(f"- `{f}`")
    else:
        lines.append("- No directories excluded (no matching skip patterns found)")
    lines.extend([
        "",
        "### What this means",
        "This analysis covers application source code only. Hazards arising from",
        "infrastructure, deployment configuration, network architecture, or",
        "operational procedures are not detected by this skill. A complete hazard",
        "analysis per ISO 14971:2019 requires consideration of all sources of harm,",
        "not just application code.",
        "",
        "False negatives are possible — the heuristic approach may miss hazardous",
        "code paths that don't match the configured patterns. This candidate list",
        "is an input to a human-led hazard analysis, not a substitute for one.",
        "",
    ])

    # Hazard Framing Disclaimer (mandatory per spec)
    lines.extend([
        "## Hazard Framing Disclaimer",
        "",
        "> These are candidate hazards inferred from code structure. Hazard analysis",
        "> requires clinical context, use environment understanding, and patient",
        "> population awareness that this skill cannot provide. Treat output as input",
        "> to a human-led hazard analysis session, not as a hazard analysis.",
        "",
    ])

    lines.extend([
        "## Recommended Sequence",
        "1. **Clinical review** — Have clinical SME review each candidate for relevance",
        "2. **Disposition** — Accept, modify, or reject each candidate with rationale",
        "3. **Clinical rationale** — Document severity and probability assessment",
        "4. **Integration** — Merge accepted candidates into formal hazard analysis",
        "5. **Risk controls** — Identify and document risk control measures",
        "",
        "## What This Report Is Not",
        "This report identifies candidate hazards by analyzing source code structure.",
        "It does not constitute a hazard analysis, a risk assessment, or any other",
        "ISO 14971 activity. The hazard candidate list and this gap report are inputs",
        "to a human-led retrospective compliance effort, not substitutes for it.",
        "",
        "The heuristic approach may miss hazardous code paths that don't match",
        "configured patterns and may produce false positives for code that resembles",
        "safety-critical logic but isn't. Teams should review all candidates and add",
        "any hazards the heuristics missed.",
    ])

    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def slugify(name: str) -> str:
    """Convert name to filename-safe slug."""
    slug = name.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    return slug.strip("-")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Reverse-engineer hazard candidates from source code (XLSX + gap report)."
    )
    parser.add_argument(
        "--source", type=str,
        help="Path to source tree root.",
    )
    parser.add_argument(
        "--scope", type=str, default=None,
        help="Path to scope statement (optional, recommended).",
    )
    parser.add_argument(
        "--risk-file", type=str, default=None,
        help="Path to existing risk file or hazard list for cross-reference.",
    )
    parser.add_argument(
        "--heuristics", type=str, default=None,
        help="Path to domain heuristic config override (JSON).",
    )
    parser.add_argument(
        "--output-dir", type=str, default=None,
        help="Directory for output files (default: output/ in repo root).",
    )
    parser.add_argument(
        "--example", action="store_true",
        help="Generate example output (no source tree needed).",
    )
    args = parser.parse_args()

    # Load heuristic config if provided
    scanners: list[Any] | None = None
    domain_name: str | None = None
    if args.heuristics:
        heuristics_path = Path(args.heuristics)
        if heuristics_path.is_file():
            try:
                scanners, domain_name = load_heuristic_config(heuristics_path)
                print(f"Loaded heuristic config: domain={domain_name}, "
                      f"scanners={len(scanners)}")
            except (json.JSONDecodeError, KeyError, ValueError) as e:
                print(f"Error loading heuristics config: {e}", file=sys.stderr)
                sys.exit(1)
        else:
            print(f"Error: Heuristics config not found: {args.heuristics}",
                  file=sys.stderr)
            sys.exit(1)

    if args.example:
        hcs, files_analyzed, files_excluded = get_example_data()
        source_name = "example-samd-app"
    elif args.source:
        source = Path(args.source).resolve()
        if not source.is_dir():
            print(f"Error: {args.source} is not a directory.", file=sys.stderr)
            sys.exit(1)
        source_name = source.name
        hcs, files_analyzed, files_excluded = walk_source_tree(source, scanners)
        if not hcs:
            print(f"No hazard candidates detected in {args.source}.")
            print("This may indicate a codebase without safety-critical code patterns.")
            print("This is expected for non-clinical applications.")
    else:
        parser.error("Provide --source /path/to/repo or --example.")

    # Deduplicate and assign IDs
    hcs = deduplicate_hcs(hcs)
    for i, hc in enumerate(hcs, 1):
        hc.hc_id = f"HC-{i:03d}"

    # Risk file cross-reference
    unmapped_hazards: list[dict[str, str]] = []
    if args.risk_file:
        risk_path = Path(args.risk_file)
        if risk_path.is_file():
            unmapped_hazards = cross_reference_risk_file(hcs, risk_path)
        else:
            print(f"WARNING: Risk file not found: {args.risk_file}", file=sys.stderr)

    if not args.scope and not args.example:
        print("WARNING: No scope statement provided. Gap report will be flagged by "
              "qa-reviewer as lacking procedural authorization.")

    if domain_name:
        print(f"Domain: {domain_name}")

    # Generate XLSX
    wb = Workbook()
    build_document_control(wb, source_name)
    build_hazard_candidates_sheet(wb, hcs)
    wb.active = 0

    # TODO: Output dir resolution is duplicated in 3 scripts. Extract to a shared
    # module if/when these scripts become a package.
    # See also: code-to-design-inputs/scripts/generate_design_inputs.py
    # See also: code-to-soup-register/scripts/generate_soup_register.py
    if args.output_dir:
        output_dir = Path(args.output_dir).resolve()
    else:
        script_dir = Path(__file__).parent
        output_dir = script_dir.parent.parent.parent / "output"
    output_dir.mkdir(exist_ok=True)

    slug = slugify(source_name)
    xlsx_path = output_dir / f"hazard-candidates-{slug}.xlsx"
    wb.save(str(xlsx_path))

    # Generate gap report
    gap_report_path = output_dir / f"gap-report-hazard-candidates-{date.today().strftime('%Y-%m-%d')}.md"
    gap_report_content = generate_gap_report(
        hcs, files_analyzed, files_excluded, source_name,
        args.scope, str(xlsx_path.name),
        unmapped_hazards=unmapped_hazards,
        risk_file_path=args.risk_file,
    )
    gap_report_path.write_text(gap_report_content)

    # Summary
    print(f"Generated: {xlsx_path}")
    print(f"Gap report: {gap_report_path}")
    print(f"  Hazard candidates: {len(hcs)}")
    print(f"  Source files analyzed: {len(files_analyzed)}")
    print(f"  Directories excluded: {len(set(files_excluded))}")

    heuristic_counts: dict[str, int] = {}
    for hc in hcs:
        heuristic_counts[hc.heuristic_matched] = heuristic_counts.get(hc.heuristic_matched, 0) + 1
    for heuristic, count in sorted(heuristic_counts.items()):
        print(f"    {heuristic}: {count}")

    if args.risk_file:
        matched = sum(1 for hc in hcs if hc.existing_haz_id)
        print(f"  Matched to existing hazards: {matched}")
        print(f"  Unmapped existing hazards: {len(unmapped_hazards)}")


if __name__ == "__main__":
    main()
