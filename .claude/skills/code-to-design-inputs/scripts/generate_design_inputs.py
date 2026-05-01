#!/usr/bin/env python3
"""Reverse-engineer design inputs from source code in an existing codebase.

Walks the source tree using 6 inclusion heuristics to identify design-input-worthy
code boundaries. Produces an IEC 62304 Clause 5.2 compliant traceability matrix
XLSX with companion gap report.

Only uses openpyxl + stdlib (Python 3.11+).

Usage:
    python scripts/generate_design_inputs.py --source /path/to/repo
    python scripts/generate_design_inputs.py --source /path/to/repo --prd /path/to/prd.md
    python scripts/generate_design_inputs.py --source /path/to/repo --hazard-candidates output/hc.xlsx
    python scripts/generate_design_inputs.py --source /path/to/repo --domain-keywords references/keywords-cardiac.json
    python scripts/generate_design_inputs.py --example
"""

import argparse
import json
import os
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants (matches design-controls / code-to-soup-register pattern)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
YELLOW_FONT = Font(color="9C6500")
ORANGE_FONT = Font(color="C65911")

# Directories to skip during source tree walk
SKIP_DIRS = {
    "node_modules", ".venv", "venv", "vendor", "dist", "build", "target",
    ".git", "__pycache__", ".tox", ".mypy_cache", ".pytest_cache",
    "test", "tests", "__tests__", ".github", "docs",
}

# File extensions to scan
SOURCE_EXTENSIONS = {
    ".py", ".js", ".ts", ".jsx", ".tsx", ".go", ".java", ".kt", ".kts",
    ".rb", ".rs", ".c", ".cpp", ".h", ".hpp", ".cs",
}


# ---------------------------------------------------------------------------
# Design Input data model
# ---------------------------------------------------------------------------
class DesignInput:
    """A single candidate design input extracted from code."""

    def __init__(
        self,
        di_id: str,
        description: str,
        source_file: str,
        source_line: int,
        di_type: str,
        heuristic: str,
        rationale: str = "GAP",
        rationale_source: str = "GAP",
        sw_safety_class: str = "GAP",
        priority: str = "P3",
    ):
        self.di_id = di_id
        self.description = description
        self.source_file = source_file
        self.source_line = source_line
        self.di_type = di_type
        self.heuristic = heuristic
        self.rationale = rationale
        self.rationale_source = rationale_source
        self.sw_safety_class = sw_safety_class
        self.priority = priority

    @property
    def source_citation(self) -> str:
        return f"{self.source_file}:{self.source_line}"


# ---------------------------------------------------------------------------
# Heuristic scanner functions
# ---------------------------------------------------------------------------

# -- Clinical keyword set for classifying Safety vs Functional --
_DEFAULT_CLINICAL_KEYWORDS = {
    "spo2", "sao2", "heart_rate", "hr", "bp", "blood_pressure",
    "temperature", "temp", "dose", "dosing", "weight", "bmi",
    "alarm", "alert", "critical", "apgar", "score", "severity",
    "patient", "clinician", "clinical", "threshold", "vital",
    "oxygen", "saturation", "respiratory", "pulse", "ecg", "ekg",
}


def load_domain_keywords(config_path: Path) -> tuple[set[str], str]:
    """Load domain keywords from JSON config.

    Returns (keyword_set, domain_name).
    Raises ValueError if the config structure is invalid.
    """
    with open(config_path) as f:
        config = json.load(f)
    if "keywords" not in config:
        raise ValueError("Domain keywords config missing required key: 'keywords'")
    kw_list = config["keywords"]
    if not isinstance(kw_list, list) or not all(isinstance(k, str) for k in kw_list):
        raise ValueError("'keywords' must be a list of strings")
    return set(kw_list), config.get("domain", "custom")


def _has_clinical_context(line: str, keywords: set[str] = _DEFAULT_CLINICAL_KEYWORDS) -> bool:
    """Check if a line contains clinical terminology."""
    lower = line.lower()
    return any(kw in lower for kw in keywords)


def scan_api_boundaries(content: str, filepath: str, keywords: set[str] = _DEFAULT_CLINICAL_KEYWORDS) -> list[DesignInput]:
    """Detect routes, exported functions, public interfaces."""
    results: list[DesignInput] = []
    patterns = [
        # Flask/FastAPI route decorators
        (r'@(?:app|router|blueprint)\.(get|post|put|delete|patch|route)\s*\(\s*["\']([^"\']+)',
         lambda m: f"System shall provide {m.group(1).upper()} endpoint at {m.group(2)}"),
        # Express.js route handlers
        (r'(?:app|router)\.(get|post|put|delete|patch)\s*\(\s*["\']([^"\']+)',
         lambda m: f"System shall handle {m.group(1).upper()} request at {m.group(2)}"),
        # Django URL patterns
        (r'path\s*\(\s*["\']([^"\']+)',
         lambda m: f"System shall provide URL endpoint at /{m.group(1)}"),
        # REST endpoint annotations (Java/Kotlin)
        (r'@(Get|Post|Put|Delete|Patch|Request)Mapping\s*\(\s*(?:value\s*=\s*)?["\']?([^"\')\s]*)',
         lambda m: f"System shall provide {m.group(1).replace('Mapping', '').upper()} endpoint at {m.group(2)}"),
    ]

    for line_num, line in enumerate(content.splitlines(), 1):
        for pattern, desc_fn in patterns:
            match = re.search(pattern, line)
            if match:
                di_type = "Interface" if _has_clinical_context(line, keywords) else "Functional"
                results.append(DesignInput(
                    di_id="",
                    description=desc_fn(match),
                    source_file=filepath,
                    source_line=line_num,
                    di_type=di_type,
                    heuristic="api_boundaries",
                ))
    return results


def scan_configuration_surfaces(content: str, filepath: str, keywords: set[str] = _DEFAULT_CLINICAL_KEYWORDS) -> list[DesignInput]:
    """Detect environment variables, feature flags, threshold constants."""
    results: list[DesignInput] = []
    patterns = [
        # Environment variable reads (Python)
        (r'os\.environ(?:\[|\.get\()[\s]*["\']([^"\']+)',
         lambda m: f"System shall use configurable parameter: {m.group(1)}"),
        # Environment variable reads (JS/TS)
        (r'process\.env\.([A-Z_]+)',
         lambda m: f"System shall use configurable parameter: {m.group(1)}"),
        # Uppercased constants with assignment (Python/config)
        (r'^([A-Z][A-Z_]{2,})\s*[:=]\s*(.+?)(?:\s*#.*)?$',
         lambda m: f"System shall define configuration constant: {m.group(1)} = {m.group(2).strip()[:60]}"),
    ]

    for line_num, line in enumerate(content.splitlines(), 1):
        stripped = line.strip()
        for pattern, desc_fn in patterns:
            match = re.search(pattern, stripped)
            if match:
                if _has_clinical_context(stripped, keywords):
                    di_type = "Safety"
                elif any(kw in match.group(1).lower() for kw in ("timeout", "rate", "latency", "max", "min", "limit")):
                    di_type = "Performance"
                else:
                    di_type = "Functional"
                results.append(DesignInput(
                    di_id="",
                    description=desc_fn(match),
                    source_file=filepath,
                    source_line=line_num,
                    di_type=di_type,
                    heuristic="configuration_surfaces",
                ))
    return results


def scan_integration_points(content: str, filepath: str) -> list[DesignInput]:
    """Detect external API calls, EHR writes, device interfaces, FHIR endpoints."""
    results: list[DesignInput] = []
    patterns = [
        # HTTP client calls (Python requests)
        (r'requests\.(get|post|put|delete|patch)\s*\(\s*(?:f?["\'])?([^"\')\s]*)',
         lambda m: f"System shall make {m.group(1).upper()} request to external service"),
        # Fetch API (JS/TS)
        (r'fetch\s*\(\s*(?:f?[`"\'])?([^`"\')\s]*)',
         lambda m: f"System shall make HTTP request to external endpoint"),
        # FHIR/HL7 module usage (import or module path — not bare keyword)
        (r'(?:from|import)\s+(?:fhir|hl7)|(?:fhir|hl7)[\._]\w+',
         lambda m: f"System shall interface with FHIR/HL7 module"),
        # FHIR resource instantiation (requires constructor call)
        # Patient included deliberately — writing patient demographics is a design input
        # Order: Bundle|Patient|Observation|DiagnosticReport|MedicationRequest (canonical)
        (r'(?:Bundle|Patient|Observation|DiagnosticReport|MedicationRequest)\s*\(',
         lambda m: f"System shall interface with FHIR/HL7 resource"),
        # Database write operations (require db/model/session/cursor/collection context)
        (r'(?:db|model|session|cursor|collection|repo|store|table)\.(save|create|insert|update|delete|execute)\s*\(',
         lambda m: f"System shall perform {m.group(1)} operation on data store"),
        # Message queue / event bus
        (r'(?:queue|broker|bus|channel|producer|kafka|rabbitmq|sns|sqs)\.(publish|send_message|emit|dispatch)\s*\(',
         lambda m: f"System shall {m.group(1)} event/message to external system"),
        # Serial / device interface
        (r'(?:serial|usb|bluetooth|ble)\.',
         lambda m: f"System shall interface with physical device"),
    ]

    for line_num, line in enumerate(content.splitlines(), 1):
        for pattern, desc_fn in patterns:
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                results.append(DesignInput(
                    di_id="",
                    description=desc_fn(match),
                    source_file=filepath,
                    source_line=line_num,
                    di_type="Interface",
                    heuristic="integration_points",
                ))
    return results


def scan_clinical_thresholds(content: str, filepath: str) -> list[DesignInput]:
    """Detect alarm limits, dosing calculations, classification boundaries."""
    results: list[DesignInput] = []
    patterns = [
        # Numeric comparisons with clinical variable names
        (r'(spo2|sao2|heart_rate|hr|bp|temperature|temp|dose|weight|bmi|score|risk|severity|alarm_threshold|alarm_limit|critical_threshold)\s*[<>=!]+\s*(\d+)',
         lambda m: f"System shall apply threshold: {m.group(1)} compared against {m.group(2)}"),
        # Alarm/threshold constants
        (r'(ALARM|THRESHOLD|CRITICAL|WARNING_LEVEL|ALERT|LIMIT|BOUNDARY|CUTOFF)[_A-Z]*\s*[=:]\s*(\d+\.?\d*)',
         lambda m: f"System shall define {m.group(1).lower().replace('_', ' ')} value: {m.group(2)}"),
        # Dosing/calculation function definitions
        (r'def\s+(calculate_\w+|compute_\w+|dose_\w+|score_\w+|classify_\w+|evaluate_\w+|assess_\w+)',
         lambda m: f"System shall implement calculation: {m.group(1).replace('_', ' ')}"),
    ]

    for line_num, line in enumerate(content.splitlines(), 1):
        for pattern, desc_fn in patterns:
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                results.append(DesignInput(
                    di_id="",
                    description=desc_fn(match),
                    source_file=filepath,
                    source_line=line_num,
                    di_type="Safety",
                    heuristic="clinical_thresholds",
                ))
    return results


def scan_data_retention(content: str, filepath: str) -> list[DesignInput]:
    """Detect storage duration policies, purge logic, data lifecycle."""
    results: list[DesignInput] = []
    patterns = [
        # Retention/expiration constants
        (r'(RETENTION|EXPIR|TTL|PURGE|ARCHIVE|DELETE_AFTER|CLEANUP|MAX_AGE|KEEP_DAYS|RETENTION_DAYS)[_A-Z]*\s*[=:]\s*(.+?)(?:\s*#.*)?$',
         lambda m: f"System shall enforce data lifecycle policy: {m.group(1).lower().replace('_', ' ')} = {m.group(2).strip()[:40]}"),
        # Time-based deletion logic
        (r'(days_old|max_age|expire_at|retention_days|keep_days)\s*[<>=]+',
         lambda m: f"System shall enforce data retention rule based on {m.group(1)}"),
        # GDPR/HIPAA deletion
        (r'(?:gdpr|hipaa|phi|pii).*(?:delete|purge|erase|remove)',
         lambda m: f"System shall implement regulatory data deletion requirement"),
    ]

    for line_num, line in enumerate(content.splitlines(), 1):
        stripped = line.strip()
        for pattern, desc_fn in patterns:
            match = re.search(pattern, stripped, re.IGNORECASE)
            if match:
                results.append(DesignInput(
                    di_id="",
                    description=desc_fn(match),
                    source_file=filepath,
                    source_line=line_num,
                    di_type="Functional",
                    heuristic="data_retention",
                ))
    return results


def scan_error_messages(content: str, filepath: str, keywords: set[str] = _DEFAULT_CLINICAL_KEYWORDS) -> list[DesignInput]:
    """Detect user-facing error messages — information-for-safety under risk controls."""
    results: list[DesignInput] = []
    patterns = [
        # Flash/toast/alert with string (Python)
        (r'(?:flash|toast|alert|notify|message)\s*\(\s*["\']([^"\']{10,})',
         lambda m: f"System shall display user message: \"{m.group(1)[:60]}...\"" if len(m.group(1)) > 60 else f"System shall display user message: \"{m.group(1)}\""),
        # HTTP error responses
        (r'(?:abort|HTTPException|HttpResponse|JsonResponse)\s*\(.*?(?:message|detail)\s*=\s*["\']([^"\']+)',
         lambda m: f"System shall return error response: \"{m.group(1)[:60]}\""),
        # JS error display
        (r'(?:setError|showError|displayError|errorMessage|alertText)\s*\(\s*[`"\']([^`"\']{10,})',
         lambda m: f"System shall display error: \"{m.group(1)[:60]}\""),
    ]

    for line_num, line in enumerate(content.splitlines(), 1):
        for pattern, desc_fn in patterns:
            match = re.search(pattern, line)
            if match:
                di_type = "Safety" if _has_clinical_context(line, keywords) else "Functional"
                results.append(DesignInput(
                    di_id="",
                    description=desc_fn(match),
                    source_file=filepath,
                    source_line=line_num,
                    di_type=di_type,
                    heuristic="error_messages",
                ))
    return results


# ---------------------------------------------------------------------------
# Source tree walker
# ---------------------------------------------------------------------------
ALL_SCANNERS = [
    scan_api_boundaries,
    scan_configuration_surfaces,
    scan_integration_points,
    scan_clinical_thresholds,
    scan_data_retention,
    scan_error_messages,
]


def walk_source_tree(
    source: Path,
    scope: str | None = None,
    keywords: set[str] | None = None,
) -> tuple[list[DesignInput], list[str]]:
    """Walk source tree and run all scanners on each file.

    Returns (design_inputs, files_analyzed).
    """
    kw = keywords if keywords is not None else _DEFAULT_CLINICAL_KEYWORDS
    all_dis: list[DesignInput] = []
    files_analyzed: list[str] = []

    for root_dir, dirs, files in os.walk(source):
        # Prune skip dirs
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]

        for fname in files:
            fpath = Path(root_dir) / fname
            ext = fpath.suffix.lower()
            if ext not in SOURCE_EXTENSIONS:
                continue

            rel_path = str(fpath.relative_to(source))
            files_analyzed.append(rel_path)

            try:
                content = fpath.read_text(errors="replace")
            except (OSError, UnicodeDecodeError):
                continue

            for scanner in ALL_SCANNERS:
                # Scanners that accept keywords get them; others don't
                try:
                    results = scanner(content, rel_path, keywords=kw)
                except TypeError:
                    results = scanner(content, rel_path)
                all_dis.extend(results)

    return all_dis, files_analyzed


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------
def deduplicate_dis(dis: list[DesignInput]) -> list[DesignInput]:
    """Remove duplicate DIs based on source location + heuristic."""
    seen: set[str] = set()
    unique: list[DesignInput] = []
    for di in dis:
        key = f"{di.source_file}:{di.source_line}:{di.heuristic}"
        if key not in seen:
            seen.add(key)
            unique.append(di)
    # Sort by file then line number
    unique.sort(key=lambda d: (d.source_file, d.source_line))
    return unique


# ---------------------------------------------------------------------------
# PRD cross-reference
# ---------------------------------------------------------------------------
def _extract_prd_statements(prd_path: Path) -> list[tuple[int, str]]:
    """Extract requirement-like statements from a PRD markdown file.

    Returns list of (line_number, statement_text).
    """
    statements: list[tuple[int, str]] = []
    content = prd_path.read_text(errors="replace")
    for line_num, line in enumerate(content.splitlines(), 1):
        stripped = line.strip()
        # Bullet points
        if re.match(r'^[-*•]\s+', stripped):
            statements.append((line_num, stripped.lstrip("-*• ")))
        # Numbered items
        elif re.match(r'^\d+\.\s+', stripped):
            statements.append((line_num, re.sub(r'^\d+\.\s+', '', stripped)))
        # Lines with requirement keywords
        elif any(kw in stripped.lower() for kw in ("shall", "must", "should", "will")):
            statements.append((line_num, stripped))
        # Lines with requirement IDs
        elif re.search(r'(REQ|FR|NFR|SR|UR)-\d+', stripped, re.IGNORECASE):
            statements.append((line_num, stripped))
    return statements


def _tokenize(text: str) -> set[str]:
    """Extract non-trivial tokens from text."""
    stop_words = {
        "the", "a", "an", "is", "are", "was", "were", "be", "been", "being",
        "have", "has", "had", "do", "does", "did", "will", "would", "could",
        "should", "shall", "must", "may", "might", "can", "to", "of", "in",
        "for", "on", "with", "at", "by", "from", "as", "into", "through",
        "and", "or", "but", "not", "no", "if", "then", "else", "when",
        "this", "that", "these", "those", "it", "its", "system",
    }
    words = re.findall(r'[a-z][a-z_]{2,}', text.lower())
    return {w for w in words if w not in stop_words}


def cross_reference_prd(
    dis: list[DesignInput], prd_path: Path
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Cross-reference DIs against PRD statements.

    Returns (doc_gaps, impl_gaps):
    - doc_gaps: code exists but PRD doesn't mention it
    - impl_gaps: PRD describes behavior that code doesn't implement
    """
    prd_statements = _extract_prd_statements(prd_path)
    if not prd_statements:
        return [], []

    # Tokenize all DIs and PRD statements
    di_tokens = [(di, _tokenize(di.description)) for di in dis]
    prd_tokens = [(line_num, stmt, _tokenize(stmt)) for line_num, stmt in prd_statements]

    matched_dis: set[int] = set()
    matched_prds: set[int] = set()

    for di_idx, (di, di_toks) in enumerate(di_tokens):
        for prd_idx, (prd_line, prd_stmt, prd_toks) in enumerate(prd_tokens):
            overlap = di_toks & prd_toks
            if len(overlap) >= 3:
                matched_dis.add(di_idx)
                matched_prds.add(prd_idx)

    # Documentation gaps: DIs not matched to any PRD statement
    doc_gaps = []
    for idx, (di, _) in enumerate(di_tokens):
        if idx not in matched_dis:
            doc_gaps.append({
                "di_id": di.di_id,
                "description": di.description,
                "source": di.source_citation,
                "gap_type": "Documentation Gap",
                "detail": "Code behavior exists but PRD does not mention it",
            })

    # Implementation gaps: PRD statements not matched to any DI
    impl_gaps = []
    for idx, (prd_line, prd_stmt, _) in enumerate(prd_tokens):
        if idx not in matched_prds:
            impl_gaps.append({
                "prd_line": str(prd_line),
                "statement": prd_stmt[:120],
                "gap_type": "Implementation Gap",
                "detail": "PRD describes behavior that code does not implement",
            })

    return doc_gaps, impl_gaps


# ---------------------------------------------------------------------------
# Priority tier assignment
# ---------------------------------------------------------------------------
def _parse_hazard_candidate_regions(hc_path: Path) -> list[tuple[str, int, int]]:
    """Parse hazard candidates XLSX to extract code regions.

    Returns list of (file, start_line, end_line).
    """
    regions: list[tuple[str, int, int]] = []
    try:
        wb = load_workbook(str(hc_path), read_only=True, data_only=True)
    except Exception:
        return regions

    # Find the sheet with hazard candidates
    sheet_name = None
    for name in wb.sheetnames:
        if "hazard" in name.lower() or "candidate" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        # Try first non-Document_Control sheet
        for name in wb.sheetnames:
            if name != "Document_Control":
                sheet_name = name
                break
    if not sheet_name:
        wb.close()
        return regions

    ws = wb[sheet_name]
    # Find Code Region column
    code_region_col = None
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
        if cell.value and "code region" in str(cell.value).lower():
            code_region_col = col_idx
            break

    if code_region_col:
        for row in ws.iter_rows(min_row=2):
            cell_val = row[code_region_col - 1].value
            if cell_val:
                # Parse "file:start-end" format
                match = re.match(r'(.+):(\d+)-(\d+)', str(cell_val))
                if match:
                    regions.append((
                        match.group(1),
                        int(match.group(2)),
                        int(match.group(3)),
                    ))

    wb.close()
    return regions


def assign_priority_tiers(
    dis: list[DesignInput], hc_path: Path | None = None
) -> None:
    """Assign P1/P2/P3 priority tiers to design inputs.

    P1: DI source overlaps hazard candidate code region (safety-critical)
    P2: DI type is Safety or heuristic is clinical_thresholds (clinical)
    P3: everything else (functional)
    """
    hc_regions: list[tuple[str, int, int]] = []
    if hc_path:
        hc_regions = _parse_hazard_candidate_regions(hc_path)

    for di in dis:
        # P1: overlaps hazard candidate region
        if hc_regions:
            for hc_file, hc_start, hc_end in hc_regions:
                if di.source_file == hc_file and hc_start <= di.source_line <= hc_end:
                    di.priority = "P1"
                    break
            else:
                # No overlap — check P2
                if di.di_type == "Safety" or di.heuristic == "clinical_thresholds":
                    di.priority = "P2"
                else:
                    di.priority = "P3"
        else:
            # No hazard candidates file — P2/P3 only
            if di.di_type == "Safety" or di.heuristic == "clinical_thresholds":
                di.priority = "P2"
            else:
                di.priority = "P3"


# ---------------------------------------------------------------------------
# Example data
# ---------------------------------------------------------------------------
def get_example_data() -> tuple[list[DesignInput], list[str]]:
    """Return example design input data for --example mode."""
    dis = [
        DesignInput("", "System shall provide GET endpoint at /api/v1/patients/<id>/vitals",
                     "app.py", 15, "Functional", "api_boundaries"),
        DesignInput("", "System shall provide POST endpoint at /api/v1/alerts",
                     "app.py", 42, "Functional", "api_boundaries"),
        DesignInput("", "System shall use configurable parameter: SPO2_LOW_THRESHOLD",
                     "config.py", 8, "Safety", "configuration_surfaces"),
        DesignInput("", "System shall use configurable parameter: ALERT_TIMEOUT_SECONDS",
                     "config.py", 12, "Performance", "configuration_surfaces"),
        DesignInput("", "System shall interface with FHIR/HL7 resource",
                     "integrations/ehr.py", 34, "Interface", "integration_points"),
        DesignInput("", "System shall make POST request to external service",
                     "integrations/ehr.py", 67, "Interface", "integration_points"),
        DesignInput("", "System shall apply threshold: spo2 compared against 88",
                     "monitor/alarms.py", 23, "Safety", "clinical_thresholds"),
        DesignInput("", "System shall define critical threshold value: 60",
                     "monitor/alarms.py", 45, "Safety", "clinical_thresholds"),
        DesignInput("", "System shall implement calculation: calculate_dose",
                     "clinical/dosing.py", 12, "Safety", "clinical_thresholds"),
        DesignInput("", "System shall enforce data lifecycle policy: retention days = 2555",
                     "storage/cleanup.py", 8, "Functional", "data_retention"),
        DesignInput("", 'System shall display user message: "WARNING: Sensor disconnected"',
                     "ui/alerts.py", 33, "Safety", "error_messages"),
        DesignInput("", 'System shall display user message: "Data export complete"',
                     "ui/notifications.py", 18, "Functional", "error_messages"),
    ]
    files = [
        "app.py", "config.py", "integrations/ehr.py",
        "monitor/alarms.py", "clinical/dosing.py",
        "storage/cleanup.py", "ui/alerts.py", "ui/notifications.py",
    ]
    return dis, files


# ---------------------------------------------------------------------------
# XLSX generation helpers
# ---------------------------------------------------------------------------
def style_header_row(ws: Any, headers: list[str]) -> None:
    """Write and style the header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws: Any, min_width: int = 12, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                line_max = max(len(line) for line in lines)
                max_len = max(max_len, line_max)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def style_data_rows(ws: Any, start_row: int, end_row: int, num_cols: int) -> None:
    """Apply alignment and borders to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# XLSX builder
# ---------------------------------------------------------------------------
def build_document_control(wb: Workbook, source_name: str) -> None:
    """Build Document_Control sheet."""
    ws = wb.active
    ws.title = "Document_Control"

    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="Document Control")
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = HEADER_ALIGN
    title_cell.border = THIN_BORDER
    for col in range(2, 5):
        ws.cell(row=1, column=col).border = THIN_BORDER
        ws.cell(row=1, column=col).fill = HEADER_FILL

    metadata = [
        ("Document Title", f"Design Input Traceability Matrix — {source_name}"),
        ("Document Version", "1.0 (draft — generated by code-to-design-inputs)"),
        ("Date Generated", date.today().strftime("%Y-%m-%d")),
        ("Source Tree", source_name),
        ("Standard Reference", "IEC 62304:2006+A1:2015 Clause 5.2"),
    ]
    for idx, (label, value) in enumerate(metadata):
        row = 3 + idx
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = CELL_ALIGN
        label_cell.border = THIN_BORDER
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.alignment = CELL_ALIGN
        value_cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def build_design_inputs_sheet(
    wb: Workbook,
    dis: list[DesignInput],
    include_priority: bool = False,
) -> None:
    """Build the Design_Inputs sheet."""
    ws = wb.create_sheet("Design_Inputs")
    headers = [
        "UN ID", "DI ID", "Description", "Source (code)", "Type",
        "Rationale", "Rationale Source", "SW Safety Class", "Gap Status",
    ]
    if include_priority:
        headers.insert(5, "Priority")

    style_header_row(ws, headers)

    for i, di in enumerate(dis, 2):
        col = 1
        ws.cell(row=i, column=col, value="GAP"); col += 1  # UN ID
        ws.cell(row=i, column=col, value=di.di_id); col += 1
        ws.cell(row=i, column=col, value=di.description); col += 1
        ws.cell(row=i, column=col, value=di.source_citation); col += 1
        ws.cell(row=i, column=col, value=di.di_type); col += 1
        if include_priority:
            ws.cell(row=i, column=col, value=di.priority); col += 1
        ws.cell(row=i, column=col, value=di.rationale); col += 1
        ws.cell(row=i, column=col, value=di.rationale_source); col += 1
        ws.cell(row=i, column=col, value=di.sw_safety_class); col += 1

        # Gap Status formula — columns shift if priority is included
        rat_col = 7 if include_priority else 6
        rat_src_col = 8 if include_priority else 7
        safety_col = 9 if include_priority else 8
        gap_col = 10 if include_priority else 9

        rat_letter = get_column_letter(rat_col)
        src_letter = get_column_letter(rat_src_col)
        saf_letter = get_column_letter(safety_col)

        ws.cell(
            row=i, column=gap_col,
            value=f'=IF(OR({rat_letter}{i}="GAP",{src_letter}{i}="GAP",{saf_letter}{i}="GAP"),"GAP","COMPLETE")',
        )

    end_row = len(dis) + 1
    num_cols = len(headers)
    if end_row >= 2:
        style_data_rows(ws, 2, end_row, num_cols)

    # Conditional formatting — only if there are data rows
    if end_row >= 2:
        gap_col_letter = get_column_letter(num_cols)
        status_range = f"{gap_col_letter}2:{gap_col_letter}{end_row}"
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=['"COMPLETE"'], fill=GREEN_FILL, font=GREEN_FONT),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=['"GAP"'], fill=RED_FILL, font=RED_FONT),
        )

        # Conditional formatting on Priority column if present
        if include_priority:
            pri_col_letter = get_column_letter(6)
            pri_range = f"{pri_col_letter}2:{pri_col_letter}{end_row}"
            ws.conditional_formatting.add(
                pri_range,
                CellIsRule(operator="equal", formula=['"P1"'], fill=RED_FILL, font=RED_FONT),
            )
            ws.conditional_formatting.add(
                pri_range,
                CellIsRule(operator="equal", formula=['"P2"'], fill=ORANGE_FILL, font=ORANGE_FONT),
            )
            ws.conditional_formatting.add(
                pri_range,
                CellIsRule(operator="equal", formula=['"P3"'], fill=GREEN_FILL, font=GREEN_FONT),
            )

    auto_width(ws)


# ---------------------------------------------------------------------------
# Gap report generation
# ---------------------------------------------------------------------------
def generate_gap_report(
    dis: list[DesignInput],
    files_analyzed: list[str],
    source_name: str,
    scope_path: str | None,
    xlsx_path: str,
    doc_gaps: list[dict[str, str]] | None = None,
    impl_gaps: list[dict[str, str]] | None = None,
    prd_path: str | None = None,
    include_priority: bool = False,
) -> str:
    """Generate companion gap report markdown."""
    today = date.today().strftime("%Y-%m-%d")
    total = len(dis)
    gap_count = total  # All rows have GAP in rationale/rationale_source/safety_class
    complete_count = 0
    scope_ref = scope_path or "none provided"
    auth_ref = f"See {scope_path}" if scope_path else "none provided"

    # Count by type
    type_counts: dict[str, int] = {}
    for di in dis:
        type_counts[di.di_type] = type_counts.get(di.di_type, 0) + 1

    # Count by heuristic
    heuristic_counts: dict[str, int] = {}
    for di in dis:
        heuristic_counts[di.heuristic] = heuristic_counts.get(di.heuristic, 0) + 1

    lines = [
        "---",
        "type: gap-report",
        "status: draft",
        f'owner: "@[github-handle]"',
        f"generated-on: {today}",
        "generated-by: code-to-design-inputs",
        f"scope-statement: {scope_ref}",
        f"source-tree: {source_name}",
        f"companion-artifact: {xlsx_path}",
        f"authorization: {auth_ref}",
        "---",
        "",
        "# Gap Report — Design Inputs (Retrospective)",
        "",
        "## Summary",
        f"- Design inputs identified: {total}",
        f"- Complete: {complete_count} (0%)",
        f"- Gaps: {gap_count} ({100 if total else 0}%)",
        f"- Source files analyzed: {len(files_analyzed)}",
        f"- Source-code traceability: {total}/{total} (100%)",
        "",
    ]

    # Priority breakdown if enabled
    if include_priority:
        p1 = sum(1 for di in dis if di.priority == "P1")
        p2 = sum(1 for di in dis if di.priority == "P2")
        p3 = sum(1 for di in dis if di.priority == "P3")
        lines.extend([
            "### Priority Breakdown",
            f"- **P1 (safety-critical):** {p1} — DI source overlaps hazard candidate code region",
            f"- **P2 (clinical):** {p2} — Safety type or clinical threshold heuristic",
            f"- **P3 (functional):** {p3} — All other design inputs",
            "",
        ])

    lines.extend([
        "### By Type",
    ])
    for di_type, count in sorted(type_counts.items()):
        lines.append(f"- {di_type}: {count}")
    lines.append("")

    lines.extend([
        "### By Heuristic",
    ])
    for heuristic, count in sorted(heuristic_counts.items()):
        lines.append(f"- {heuristic}: {count}")
    lines.append("")

    # Files analyzed
    lines.extend([
        "## Source Files Analyzed",
        "",
    ])
    for f in sorted(files_analyzed):
        lines.append(f"- `{f}`")
    lines.append("")

    # Gaps by category
    lines.extend([
        "## Gaps by Category",
        "",
        f"### Rationale Required ({total})",
        "Design rationale not documented for any identified design input.",
        "Owner: Engineering / Product.",
        "",
        "| DI ID | Description | Source | Type | Action |",
        "|-------|-------------|--------|------|--------|",
    ])
    for di in dis:
        lines.append(f"| {di.di_id} | {di.description[:60]} | {di.source_citation} | {di.di_type} | Document rationale |")

    lines.extend([
        "",
        f"### SW Safety Class Required ({total})",
        "Software safety classification (A/B/C) not assigned for any design input.",
        "Owner: Engineering / Regulatory Affairs.",
        "",
    ])

    # PRD cross-reference results
    doc_gaps = doc_gaps or []
    impl_gaps = impl_gaps or []
    if prd_path:
        lines.extend([
            "## PRD Cross-Reference",
            f"PRD analyzed: `{prd_path}`",
            "",
        ])
        if doc_gaps:
            lines.extend([
                f"### Documentation Gaps ({len(doc_gaps)})",
                "Code behavior exists but PRD does not mention it.",
                "Owner: Product / Regulatory Affairs.",
                "",
                "| DI ID | Description | Source | Action |",
                "|-------|-------------|--------|--------|",
            ])
            for gap in doc_gaps:
                lines.append(f"| {gap['di_id']} | {gap['description'][:60]} | {gap['source']} | Add to PRD or document rationale for exclusion |")
            lines.append("")

        if impl_gaps:
            lines.extend([
                f"### Implementation Gaps ({len(impl_gaps)})",
                "PRD describes behavior that code does not implement.",
                "Owner: Engineering / Product.",
                "",
                "| PRD Line | Statement | Action |",
                "|----------|-----------|--------|",
            ])
            for gap in impl_gaps:
                lines.append(f"| {gap['prd_line']} | {gap['statement'][:80]} | Verify: deferred scope, bug, or intentional exclusion |")
            lines.append("")

        if not doc_gaps and not impl_gaps:
            lines.extend([
                "All discovered DIs matched to PRD statements and vice versa.",
                "Manual review still recommended — keyword matching has limitations.",
                "",
            ])

    lines.extend([
        "## Recommended Sequence",
        "1. **Safety classification** — assign SW safety class (A/B/C) based on risk analysis",
        "2. **Rationale documentation** — document why each design input exists",
        "3. **PRD alignment** — resolve documentation and implementation gaps",
        "4. **Traceability** — link DIs to user needs (UN IDs) and design outputs",
        "5. **Verification planning** — define V&V activities for each DI",
        "",
        "## What This Report Is Not",
        "This report identifies candidate design inputs by analyzing source code structure.",
        "It does not constitute a design input review, a requirements analysis, or any other",
        "regulated activity. The design input list and this gap report are inputs to a",
        "human-led retrospective compliance effort, not substitutes for it.",
        "",
        "The heuristic approach may miss design inputs that don't match configured patterns",
        "and may produce false positives for code that looks like a design input but isn't.",
        "Teams should review all candidates and add any design inputs the heuristics missed.",
    ])

    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def slugify(name: str) -> str:
    """Convert name to filename-safe slug."""
    slug = name.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    return slug.strip("-")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Reverse-engineer design inputs from source code (XLSX + gap report)."
    )
    parser.add_argument(
        "--source", type=str,
        help="Path to source tree root.",
    )
    parser.add_argument(
        "--scope", type=str, default=None,
        help="Path to scope statement (optional, recommended).",
    )
    parser.add_argument(
        "--prd", type=str, default=None,
        help="Path to PRD or intended use statement for cross-reference.",
    )
    parser.add_argument(
        "--hazard-candidates", type=str, default=None,
        help="Path to hazard candidates XLSX for priority tier assignment.",
    )
    parser.add_argument(
        "--domain-keywords", type=str, default=None,
        help="Path to domain keyword config (JSON). Replaces default clinical keywords.",
    )
    parser.add_argument(
        "--output-dir", type=str, default=None,
        help="Directory for output files (default: output/ in repo root).",
    )
    parser.add_argument(
        "--example", action="store_true",
        help="Generate example output (no source tree needed).",
    )
    args = parser.parse_args()

    # Load domain keywords if provided
    loaded_keywords: set[str] | None = None
    if args.domain_keywords:
        kw_path = Path(args.domain_keywords)
        if kw_path.is_file():
            try:
                loaded_keywords, domain = load_domain_keywords(kw_path)
                print(f"Loaded domain keywords: domain={domain}, "
                      f"keywords={len(loaded_keywords)}")
            except (json.JSONDecodeError, KeyError, ValueError) as e:
                print(f"Error loading domain keywords: {e}", file=sys.stderr)
                sys.exit(1)
        else:
            print(f"Error: Domain keywords file not found: {args.domain_keywords}",
                  file=sys.stderr)
            sys.exit(1)

    if args.example:
        dis, files_analyzed = get_example_data()
        source_name = "example-samd-app"
    elif args.source:
        source = Path(args.source).resolve()
        if not source.is_dir():
            print(f"Error: {args.source} is not a directory.", file=sys.stderr)
            sys.exit(1)
        source_name = source.name
        dis, files_analyzed = walk_source_tree(source, args.scope, keywords=loaded_keywords)
        if not dis:
            print(f"No design inputs detected in {args.source}.")
            print("This may indicate a codebase with no detectable API boundaries,")
            print("configuration surfaces, or clinical logic. Review manually.")
            # Still generate empty report — graceful handling
    else:
        parser.error("Provide --source /path/to/repo or --example.")

    # Deduplicate and assign IDs
    dis = deduplicate_dis(dis)
    for i, di in enumerate(dis, 1):
        di.di_id = f"DI-{i:03d}"

    # Priority tiers
    include_priority = args.hazard_candidates is not None
    hc_path = Path(args.hazard_candidates) if args.hazard_candidates else None
    if include_priority or any(di.di_type == "Safety" or di.heuristic == "clinical_thresholds" for di in dis):
        assign_priority_tiers(dis, hc_path)
        include_priority = True

    # PRD cross-reference
    doc_gaps: list[dict[str, str]] = []
    impl_gaps: list[dict[str, str]] = []
    if args.prd:
        prd_path = Path(args.prd)
        if prd_path.is_file():
            doc_gaps, impl_gaps = cross_reference_prd(dis, prd_path)
        else:
            print(f"WARNING: PRD file not found: {args.prd}", file=sys.stderr)

    if not args.scope and not args.example:
        print("WARNING: No scope statement provided. Gap report will be flagged by "
              "qa-reviewer as lacking procedural authorization.")

    # Generate XLSX
    wb = Workbook()
    build_document_control(wb, source_name)
    build_design_inputs_sheet(wb, dis, include_priority=include_priority)
    wb.active = 0

    # TODO: Output dir resolution is duplicated in 3 scripts. Extract to a shared
    # module if/when these scripts become a package.
    # See also: code-to-hazard-candidates/scripts/generate_hazard_candidates.py
    # See also: code-to-soup-register/scripts/generate_soup_register.py
    if args.output_dir:
        output_dir = Path(args.output_dir).resolve()
    else:
        script_dir = Path(__file__).parent
        output_dir = script_dir.parent.parent.parent / "output"
    output_dir.mkdir(exist_ok=True)

    slug = slugify(source_name)
    xlsx_path = output_dir / f"design-inputs-{slug}.xlsx"
    wb.save(str(xlsx_path))

    # Generate gap report
    gap_report_path = output_dir / f"gap-report-design-inputs-{date.today().strftime('%Y-%m-%d')}.md"
    gap_report_content = generate_gap_report(
        dis, files_analyzed, source_name, args.scope, str(xlsx_path.name),
        doc_gaps=doc_gaps, impl_gaps=impl_gaps,
        prd_path=args.prd,
        include_priority=include_priority,
    )
    gap_report_path.write_text(gap_report_content)

    # Summary
    print(f"Generated: {xlsx_path}")
    print(f"Gap report: {gap_report_path}")
    print(f"  Design inputs: {len(dis)}")
    print(f"  Source files analyzed: {len(files_analyzed)}")

    type_counts: dict[str, int] = {}
    for di in dis:
        type_counts[di.di_type] = type_counts.get(di.di_type, 0) + 1
    for di_type, count in sorted(type_counts.items()):
        print(f"    {di_type}: {count}")

    if include_priority:
        p1 = sum(1 for di in dis if di.priority == "P1")
        p2 = sum(1 for di in dis if di.priority == "P2")
        p3 = sum(1 for di in dis if di.priority == "P3")
        print(f"  Priority: P1={p1}, P2={p2}, P3={p3}")

    if doc_gaps:
        print(f"  Documentation gaps: {len(doc_gaps)}")
    if impl_gaps:
        print(f"  Implementation gaps: {len(impl_gaps)}")


if __name__ == "__main__":
    main()
